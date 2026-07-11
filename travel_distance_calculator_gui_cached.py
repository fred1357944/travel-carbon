#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
出差距離計算與地圖產生工具 (GUI版本 - 優化版 + 智能縮放)
功能：
1. 讀取出差地點資料
2. 計算實際距離（國內開車/國外飛行）
3. 產生路線地圖並截圖
4. 將結果整合到Excel
5. 快取機制避免重複計算和產生地圖
6. 智能地圖縮放根據距離調整
"""

import os
import re
import sys
import time
import json
import pickle
import requests
import pandas as pd
import numpy as np
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
try:
    from PIL import Image, ImageTk
except ImportError:
    Image = None
    ImageTk = None
import folium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from geopy.distance import geodesic
import warnings
warnings.filterwarnings('ignore')

# 座標資料庫
COORDINATES = {
    # 台灣地區（政府機關位置）
    '輔仁大學': {'lat': 25.0356, 'lon': 121.4320, 'name': '輔仁大學正門', 'type': 'start'},
    '台北': {'lat': 25.0330, 'lon': 121.5654, 'name': '台北市政府', 'type': 'domestic'},
    '臺北': {'lat': 25.0330, 'lon': 121.5654, 'name': '台北市政府', 'type': 'domestic'},
    '新北': {'lat': 25.0119, 'lon': 121.4653, 'name': '新北市政府', 'type': 'domestic'},
    '桃園': {'lat': 24.9936, 'lon': 121.3010, 'name': '桃園市政府', 'type': 'domestic'},
    '新竹': {'lat': 24.8138, 'lon': 120.9675, 'name': '新竹市政府', 'type': 'domestic'},
    '苗栗': {'lat': 24.5602, 'lon': 120.8214, 'name': '苗栗縣政府', 'type': 'domestic'},
    '台中': {'lat': 24.1639, 'lon': 120.6478, 'name': '台中市政府', 'type': 'domestic'},
    '臺中': {'lat': 24.1639, 'lon': 120.6478, 'name': '台中市政府', 'type': 'domestic'},
    '彰化': {'lat': 24.0757, 'lon': 120.5442, 'name': '彰化縣政府', 'type': 'domestic'},
    '南投': {'lat': 23.9611, 'lon': 120.9719, 'name': '南投縣政府', 'type': 'domestic'},
    '雲林': {'lat': 23.7092, 'lon': 120.5410, 'name': '雲林縣政府', 'type': 'domestic'},
    '嘉義': {'lat': 23.4801, 'lon': 120.4538, 'name': '嘉義市政府', 'type': 'domestic'},
    '台南': {'lat': 22.9997, 'lon': 120.2269, 'name': '台南市政府', 'type': 'domestic'},
    '臺南': {'lat': 22.9997, 'lon': 120.2269, 'name': '台南市政府', 'type': 'domestic'},
    '高雄': {'lat': 22.6273, 'lon': 120.3014, 'name': '高雄市政府', 'type': 'domestic'},
    '屏東': {'lat': 22.6821, 'lon': 120.4871, 'name': '屏東縣政府', 'type': 'domestic'},
    '宜蘭': {'lat': 24.7021, 'lon': 121.7377, 'name': '宜蘭縣政府', 'type': 'domestic'},
    '花蓮': {'lat': 23.9871, 'lon': 121.6015, 'name': '花蓮縣政府', 'type': 'domestic'},
    '台東': {'lat': 22.7972, 'lon': 121.0713, 'name': '台東縣政府', 'type': 'domestic'},
    '臺東': {'lat': 22.7972, 'lon': 121.0713, 'name': '台東縣政府', 'type': 'domestic'},
    '澎湖': {'lat': 23.5711, 'lon': 119.5793, 'name': '澎湖縣政府', 'type': 'island'},
    '金門': {'lat': 24.4493, 'lon': 118.3766, 'name': '金門縣政府', 'type': 'island'},
    '馬祖': {'lat': 26.1505, 'lon': 119.9498, 'name': '連江縣政府', 'type': 'island'},
    '基隆': {'lat': 25.1324, 'lon': 121.7391, 'name': '基隆市政府', 'type': 'domestic'},
    
    # 台灣地區特殊地點
    '東海': {'lat': 24.1795, 'lon': 120.5933, 'name': '東海大學', 'type': 'domestic'},
    '東海大學': {'lat': 24.1795, 'lon': 120.5933, 'name': '東海大學', 'type': 'domestic'},
    '清華': {'lat': 24.7963, 'lon': 120.9968, 'name': '國立清華大學', 'type': 'domestic'},
    '清華大學': {'lat': 24.7963, 'lon': 120.9968, 'name': '國立清華大學', 'type': 'domestic'},
    '政大': {'lat': 24.9883, 'lon': 121.5763, 'name': '國立政治大學', 'type': 'domestic'},
    '陽明交大': {'lat': 24.7876, 'lon': 120.9973, 'name': '國立陽明交通大學', 'type': 'domestic'},
    '中山醫': {'lat': 24.1237, 'lon': 120.6508, 'name': '中山醫學大學', 'type': 'domestic'},
    '淡江': {'lat': 25.1747, 'lon': 121.4518, 'name': '淡江大學', 'type': 'domestic'},
    '淡江大學': {'lat': 25.1747, 'lon': 121.4518, 'name': '淡江大學', 'type': 'domestic'},
    '文化': {'lat': 25.1369, 'lon': 121.5303, 'name': '文化大學', 'type': 'domestic'},
    '中原大學': {'lat': 24.9570, 'lon': 121.2412, 'name': '中原大學', 'type': 'domestic'},
    '靜宜': {'lat': 24.2267, 'lon': 120.5784, 'name': '靜宜大學', 'type': 'domestic'},
    '逢甲': {'lat': 24.1778, 'lon': 120.6455, 'name': '逢甲大學', 'type': 'domestic'},
    '明道': {'lat': 23.8648, 'lon': 120.4685, 'name': '明道大學', 'type': 'domestic'},
    '東吳': {'lat': 25.1296, 'lon': 121.5135, 'name': '東吳大學', 'type': 'domestic'},
    '臺師': {'lat': 25.0261, 'lon': 121.5277, 'name': '國立臺灣師範大學', 'type': 'domestic'},
    '北醫': {'lat': 25.0266, 'lon': 121.5616, 'name': '臺北醫學大學', 'type': 'domestic'},
    '臺藝': {'lat': 25.0038, 'lon': 121.4498, 'name': '國立臺灣藝術大學', 'type': 'domestic'},
    '成功': {'lat': 22.9968, 'lon': 120.2188, 'name': '國立成功大學', 'type': 'domestic'},
    '臺灣海洋': {'lat': 25.1506, 'lon': 121.7731, 'name': '國立臺灣海洋大學', 'type': 'domestic'},
    '耕莘護': {'lat': 25.0361, 'lon': 121.4315, 'name': '耕莘健康管理專科學校', 'type': 'domestic'},
    '北一': {'lat': 25.0422, 'lon': 121.5094, 'name': '北一女中', 'type': 'domestic'},
    '教育部': {'lat': 25.0321, 'lon': 121.5192, 'name': '教育部', 'type': 'domestic'},
    '南科大': {'lat': 23.0239, 'lon': 120.2783, 'name': '南臺科技大學', 'type': 'domestic'},
    
    # 台灣地區特殊行政區
    '南港': {'lat': 25.0550, 'lon': 121.6069, 'name': '南港區', 'type': 'domestic'},
    '中和': {'lat': 24.9989, 'lon': 121.4931, 'name': '中和區', 'type': 'domestic'},
    '汐止': {'lat': 25.0627, 'lon': 121.6581, 'name': '汐止區', 'type': 'domestic'},
    '三重': {'lat': 25.0607, 'lon': 121.4884, 'name': '三重區', 'type': 'domestic'},
    '五股': {'lat': 25.0827, 'lon': 121.4377, 'name': '五股區', 'type': 'domestic'},
    '新莊': {'lat': 25.0361, 'lon': 121.4506, 'name': '新莊區', 'type': 'domestic'},
    '板橋': {'lat': 25.0115, 'lon': 121.4591, 'name': '板橋區', 'type': 'domestic'},
    '淡水': {'lat': 25.1694, 'lon': 121.4408, 'name': '淡水區', 'type': 'domestic'},
    '三峽': {'lat': 24.9340, 'lon': 121.3684, 'name': '三峽區', 'type': 'domestic'},
    '花東': {'lat': 23.9871, 'lon': 121.6015, 'name': '花東地區', 'type': 'domestic'},
    
    # 特殊名稱對應
    '集思台大': {'lat': 25.0174, 'lon': 121.5397, 'name': '集思台大會議中心', 'type': 'domestic'},
    '新光醫院': {'lat': 25.0944, 'lon': 121.5201, 'name': '新光醫院', 'type': 'domestic'},
    '台大醫院': {'lat': 25.0404, 'lon': 121.5188, 'name': '台大醫院', 'type': 'domestic'},
    
    # 機場
    '桃園機場': {'lat': 25.0777, 'lon': 121.2325, 'name': '桃園國際機場', 'type': 'airport'},
    '松山機場': {'lat': 25.0694, 'lon': 121.5526, 'name': '台北松山機場', 'type': 'airport'},
    '澎湖機場': {'lat': 23.5676, 'lon': 119.6282, 'name': '澎湖機場', 'type': 'airport'},
    '金門機場': {'lat': 24.4274, 'lon': 118.3592, 'name': '金門機場', 'type': 'airport'},
    
    # 港澳地區特殊處理
    '香港': {'lat': 22.3080, 'lon': 113.9185, 'name': '香港國際機場', 'type': 'international'},
    '澳門': {'lat': 22.1496, 'lon': 113.5910, 'name': '澳門國際機場', 'type': 'international'},
    
    # 新增特定地點（根據用戶建議）
    '泗水': {'lat': -7.3797, 'lon': 112.7913, 'name': '朱安達國際機場', 'type': 'international'},
    '浙江': {'lat': 30.2294, 'lon': 120.4344, 'name': '杭州蕭山國際機場', 'type': 'international'},
    '威尼斯': {'lat': 41.8002, 'lon': 12.2389, 'name': '羅馬菲烏米奇諾機場', 'type': 'international'},
    '羅馬': {'lat': 41.8002, 'lon': 12.2389, 'name': '羅馬菲烏米奇諾機場', 'type': 'international'},
    '山東': {'lat': 36.8577, 'lon': 117.2157, 'name': '濟南遙墻國際機場', 'type': 'international'},
    '費城': {'lat': 39.8719, 'lon': -75.2411, 'name': '費城國際機場', 'type': 'international'},
    '維也納': {'lat': 48.1103, 'lon': 16.5697, 'name': '維也納國際機場', 'type': 'international'},
    '布拉格': {'lat': 50.1008, 'lon': 14.2632, 'name': '布拉格瓦茨拉夫·哈維爾國際機場', 'type': 'international'},
    '蒙特婁': {'lat': 45.4706, 'lon': -73.7408, 'name': '蒙特婁皮埃爾·埃利奧特·特魯多國際機場', 'type': 'international'},
    '里司本': {'lat': -31.9402, 'lon': 115.9671, 'name': '珀斯機場', 'type': 'international'},
    
    # 國際主要機場
    '日本': {'lat': 35.5494, 'lon': 139.7798, 'name': '成田國際機場', 'type': 'international'},
    '韓國': {'lat': 37.4602, 'lon': 126.4407, 'name': '仁川國際機場', 'type': 'international'},
    '新加坡': {'lat': 1.3644, 'lon': 103.9915, 'name': '樟宜機場', 'type': 'international'},
    '泰國': {'lat': 13.6900, 'lon': 100.7501, 'name': '素萬那普機場', 'type': 'international'},
    '美國': {'lat': 33.9425, 'lon': -118.4081, 'name': '洛杉磯國際機場', 'type': 'international'},
    '德國': {'lat': 50.0379, 'lon': 8.5622, 'name': '法蘭克福機場', 'type': 'international'},
    '英國': {'lat': 51.4700, 'lon': -0.4543, 'name': '希斯洛機場', 'type': 'international'},
    '法國': {'lat': 49.0097, 'lon': 2.5479, 'name': '戴高樂機場', 'type': 'international'},
    '澳洲': {'lat': -33.9399, 'lon': 151.1753, 'name': '雪梨機場', 'type': 'international'},
    '加拿大': {'lat': 43.6777, 'lon': -79.6248, 'name': '多倫多機場', 'type': 'international'},
    '印度': {'lat': 28.5562, 'lon': 77.1000, 'name': '英迪拉甘地機場', 'type': 'international'},
    '巴西': {'lat': -23.4356, 'lon': -46.4731, 'name': '聖保羅國際機場', 'type': 'international'},
    '墨西哥': {'lat': 19.4363, 'lon': -99.0721, 'name': '墨西哥城機場', 'type': 'international'},
    '印尼': {'lat': -6.1256, 'lon': 106.6559, 'name': '蘇加諾哈達機場', 'type': 'international'},
    '菲律賓': {'lat': 14.5086, 'lon': 121.0198, 'name': '尼諾伊艾奎諾機場', 'type': 'international'},
    '馬來西亞': {'lat': 2.7456, 'lon': 101.7072, 'name': '吉隆坡國際機場', 'type': 'international'},
    '越南': {'lat': 10.8184, 'lon': 106.6519, 'name': '新山一國際機場', 'type': 'international'},
    '迦納': {'lat': 5.6052, 'lon': -0.1668, 'name': '科托卡國際機場', 'type': 'international'},
    '南非': {'lat': -26.1367, 'lon': 28.2460, 'name': '約翰尼斯堡機場', 'type': 'international'},
    '埃及': {'lat': 30.1219, 'lon': 31.4056, 'name': '開羅國際機場', 'type': 'international'},
    '肯亞': {'lat': -1.3192, 'lon': 36.9278, 'name': '喬莫肯雅塔機場', 'type': 'international'},
    '衣索比亞': {'lat': 8.9779, 'lon': 38.7993, 'name': '博萊國際機場', 'type': 'international'},
}

def calculate_zoom_level(distance_km):
    """根據距離計算適當的地圖縮放級別"""
    if distance_km < 10:
        return 13  # 非常近的距離（如同區域）
    elif distance_km < 20:
        return 12  # 很近的距離（如新北）
    elif distance_km < 40:
        return 11  # 近距離（如台北）
    elif distance_km < 80:
        return 10  # 中等距離
    elif distance_km < 150:
        return 9   # 較遠距離（如台中）
    elif distance_km < 300:
        return 8   # 遠距離（如高雄）
    elif distance_km < 600:
        return 7   # 很遠距離
    elif distance_km < 1500:
        return 6   # 國際短程
    elif distance_km < 3000:
        return 5   # 國際中程
    elif distance_km < 6000:
        return 4   # 國際長程
    else:
        return 3   # 跨洲長程

class TravelDistanceCalculator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("出差距離計算與地圖產生工具")
        self.root.geometry("1200x800")
        
        # 確保視窗在最前面並顯示
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(self.root.attributes, '-topmost', False)
        self.root.update()
        self.root.deiconify()
        
        # 設定樣式
        style = ttk.Style()
        style.theme_use('clam')
        
        # 變數
        self.input_file = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.path.expanduser("~/Downloads/差旅費"))
        self.start_row = tk.IntVar(value=1)
        self.end_row = tk.IntVar(value=10)
        self.progress = tk.DoubleVar()
        self.status_text = tk.StringVar(value="準備就緒")
        
        # 資料存儲
        self.df_locations = None
        self.results = []
        
        # 快取機制
        self.route_cache = {}  # 儲存已計算的路線
        self.map_cache = {}    # 儲存已產生的地圖檔案
        self.load_cache()      # 載入既有快取
        
        self.setup_ui()
        
    def load_cache(self):
        """載入快取檔案"""
        cache_dir = os.path.expanduser("~/Downloads/差旅費/.cache")
        os.makedirs(cache_dir, exist_ok=True)
        
        # 載入路線快取
        route_cache_file = os.path.join(cache_dir, 'route_cache.pkl')
        if os.path.exists(route_cache_file):
            try:
                with open(route_cache_file, 'rb') as f:
                    self.route_cache = pickle.load(f)
                print(f"載入 {len(self.route_cache)} 條快取路線")
            except:
                self.route_cache = {}
                
        # 載入地圖快取
        map_cache_file = os.path.join(cache_dir, 'map_cache.json')
        if os.path.exists(map_cache_file):
            try:
                with open(map_cache_file, 'r', encoding='utf-8') as f:
                    self.map_cache = json.load(f)
                print(f"載入 {len(self.map_cache)} 個快取地圖")
            except:
                self.map_cache = {}
                
    def save_cache(self):
        """儲存快取檔案"""
        cache_dir = os.path.expanduser("~/Downloads/差旅費/.cache")
        
        # 儲存路線快取
        route_cache_file = os.path.join(cache_dir, 'route_cache.pkl')
        with open(route_cache_file, 'wb') as f:
            pickle.dump(self.route_cache, f)
            
        # 儲存地圖快取
        map_cache_file = os.path.join(cache_dir, 'map_cache.json')
        with open(map_cache_file, 'w', encoding='utf-8') as f:
            json.dump(self.map_cache, f, ensure_ascii=False, indent=2)
        
    def setup_ui(self):
        """設置UI介面"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 檔案選擇區
        file_frame = ttk.LabelFrame(main_frame, text="檔案設定", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(file_frame, text="輸入檔案:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(file_frame, textvariable=self.input_file, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(file_frame, text="瀏覽", command=self.browse_input_file).grid(row=0, column=2)
        
        ttk.Label(file_frame, text="輸出目錄:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.output_dir, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(file_frame, text="瀏覽", command=self.browse_output_dir).grid(row=1, column=2)
        
        # 處理範圍設定
        range_frame = ttk.LabelFrame(main_frame, text="處理範圍", padding="10")
        range_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(range_frame, text="起始列:").grid(row=0, column=0, sticky=tk.W)
        ttk.Spinbox(range_frame, from_=1, to=10000, textvariable=self.start_row, width=10).grid(row=0, column=1, padx=5)
        
        ttk.Label(range_frame, text="結束列:").grid(row=0, column=2, sticky=tk.W, padx=(20, 0))
        ttk.Spinbox(range_frame, from_=1, to=10000, textvariable=self.end_row, width=10).grid(row=0, column=3, padx=5)
        
        # 快取資訊
        cache_info = ttk.Label(range_frame, text=f"快取: {len(self.route_cache)} 條路線, {len(self.map_cache)} 個地圖")
        cache_info.grid(row=0, column=4, padx=(20, 0))
        
        # 控制按鈕
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="載入資料", command=self.load_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="計算距離", command=self.calculate_distances).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="產生地圖", command=self.generate_maps).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="匯出Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="清除快取", command=self.clear_cache).pack(side=tk.LEFT, padx=5)
        
        # 進度條
        progress_frame = ttk.Frame(main_frame)
        progress_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress, maximum=100)
        self.progress_bar.pack(fill=tk.X, expand=True)
        
        # 狀態列
        status_frame = ttk.Frame(main_frame)
        status_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        ttk.Label(status_frame, textvariable=self.status_text).pack(side=tk.LEFT)
        
        # 結果顯示區
        result_frame = ttk.LabelFrame(main_frame, text="處理結果", padding="10")
        result_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # 建立Treeview
        columns = ('序號', '傳票編號', '目的地', '類型', '距離(km)', '時間', '地圖', '快取')
        self.tree = ttk.Treeview(result_frame, columns=columns, show='headings', height=15)
        
        # 設定欄位
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100 if col != '快取' else 60)
        
        # 捲軸
        scrollbar = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 配置grid權重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
    def clear_cache(self):
        """清除快取"""
        if messagebox.askyesno("確認", "確定要清除所有快取嗎？"):
            self.route_cache = {}
            self.map_cache = {}
            self.save_cache()
            messagebox.showinfo("成功", "快取已清除")
            self.status_text.set("快取已清除")
        
    def browse_input_file(self):
        """選擇輸入檔案"""
        filename = filedialog.askopenfilename(
            title="選擇Excel檔案",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.input_file.set(filename)
            
    def browse_output_dir(self):
        """選擇輸出目錄"""
        directory = filedialog.askdirectory(title="選擇輸出目錄")
        if directory:
            self.output_dir.set(directory)
            
    def load_data(self):
        """載入Excel資料"""
        if not self.input_file.get():
            messagebox.showerror("錯誤", "請先選擇輸入檔案")
            return
            
        try:
            self.status_text.set("正在載入資料...")
            self.df_locations = pd.read_excel(self.input_file.get())
            
            # 顯示資料資訊
            total_rows = len(self.df_locations)
            messagebox.showinfo("載入成功", f"成功載入 {total_rows} 筆資料")
            self.status_text.set(f"已載入 {total_rows} 筆資料")
            
            # 更新範圍設定的最大值
            self.end_row.set(min(10, total_rows))
            
        except Exception as e:
            messagebox.showerror("錯誤", f"載入資料失敗: {str(e)}")
            self.status_text.set("載入失敗")
            
    def calculate_distances(self):
        """計算距離"""
        if self.df_locations is None:
            messagebox.showerror("錯誤", "請先載入資料")
            return
            
        # 在新執行緒中執行計算
        thread = threading.Thread(target=self._calculate_distances_thread)
        thread.start()
        
    def _calculate_distances_thread(self):
        """計算距離的執行緒函數"""
        try:
            start_idx = self.start_row.get() - 1
            end_idx = self.end_row.get()
            
            # 清空結果
            self.results = []
            self.tree.delete(*self.tree.get_children())
            
            # 處理選定範圍的資料
            subset = self.df_locations.iloc[start_idx:end_idx]
            total = len(subset)
            
            cache_hit = 0
            api_calls = 0
            
            for idx, (_, row) in enumerate(subset.iterrows()):
                self.status_text.set(f"正在處理第 {idx+1}/{total} 筆...")
                self.progress.set((idx + 1) / total * 100)
                
                # 取得目的地
                destinations = str(row.get('出差地點', '')).split('、') if pd.notna(row.get('出差地點', '')) else []
                
                if destinations and destinations[0]:
                    destination = destinations[0]  # 取第一個目的地
                    
                    # 判斷是國內還是國外
                    is_domestic = self._is_domestic(destination)
                    
                    # 特殊處理離島（使用飛機）
                    if destination in ['澎湖', '金門', '馬祖'] or (destination in COORDINATES and COORDINATES[destination]['type'] == 'island'):
                        # 台灣離島：使用飛機
                        result = self._calculate_island_flight_distance(row, destination)
                        if result and result.get('cached'):
                            cache_hit += 1
                    elif is_domestic:
                        # 台灣本島：統一從輔大出發開車
                        result = self._calculate_driving_distance(row, destination)
                        if result and result.get('cached'):
                            cache_hit += 1
                        else:
                            api_calls += 1
                    else:
                        # 國外：計算飛行距離（從桃園機場出發）
                        result = self._calculate_flight_distance(row, destination)
                        if result and result.get('cached'):
                            cache_hit += 1
                    
                    if result:
                        self.results.append(result)
                        # 更新UI
                        self.root.after(0, self._add_result_to_tree, result)
                
                # 減少API請求頻率
                if api_calls > 0 and (api_calls % 5 == 0):
                    time.sleep(1)
                    
            # 儲存快取
            self.save_cache()
            
            self.status_text.set(f"計算完成！共 {len(self.results)} 筆，快取命中 {cache_hit} 筆，新計算 {api_calls} 筆")
            self.progress.set(100)
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("錯誤", f"計算失敗: {str(e)}"))
            self.status_text.set("計算失敗")
            
    def _is_domestic(self, destination):
        """判斷是否為國內地點（台灣本島用開車，離島用飛機）"""
        # 台灣本島縣市列表（統一從輔大出發開車）
        taiwan_mainland_cities = ['台北', '臺北', '新北', '桃園', '新竹', '苗栗', '台中', '臺中', 
                                '彰化', '南投', '雲林', '嘉義', '台南', '臺南', '高雄', '屏東',
                                '宜蘭', '花蓮', '台東', '臺東', '基隆']
        
        # 台灣離島（需要搭飛機）
        taiwan_islands = ['澎湖', '金門', '馬祖']
        
        # 檢查是否為台灣本島（開車）
        if any(city in destination for city in taiwan_mainland_cities):
            return True
            
        # 檢查是否為台灣離島（飛機）
        if any(island in destination for island in taiwan_islands):
            return False  # 離島視為非本島，使用飛機
            
        # 特殊地點判斷
        if destination in COORDINATES:
            coord_info = COORDINATES[destination]
            if coord_info['type'] == 'domestic':
                return True
            elif coord_info['type'] == 'island':
                return False  # 離島使用飛機
                
        return False  # 其他都視為國外
        
    def _calculate_driving_distance(self, row, destination):
        """計算開車距離"""
        if destination not in COORDINATES:
            return None
        
        # 建立路線的唯一識別碼
        route_key = f"輔仁大學-{destination}-driving"
        
        # 檢查快取
        if route_key in self.route_cache:
            cached_data = self.route_cache[route_key]
            return {
                '序號': row['序號'],
                '傳票編號': row['傳票編號'],
                '目的地': cached_data['目的地'],
                '類型': cached_data['類型'],
                '距離(km)': cached_data['距離(km)'],
                '時間': cached_data['時間'],
                '起點': cached_data['起點'],
                '終點': cached_data['終點'],
                'route': cached_data.get('route'),
                'route_key': route_key,
                'cached': True
            }
            
        try:
            # 使用OSRM API計算實際路線
            start = COORDINATES['輔仁大學']
            end = COORDINATES[destination]
            
            url = f"http://router.project-osrm.org/route/v1/driving/{start['lon']},{start['lat']};{end['lon']},{end['lat']}"
            params = {'overview': 'full', 'geometries': 'geojson'}
            
            response = requests.get(url, params=params, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                if data['code'] == 'Ok':
                    route = data['routes'][0]
                    distance_km = round(route['distance'] / 1000, 1)
                    duration_min = round(route['duration'] / 60)
                    
                    # 儲存到快取
                    self.route_cache[route_key] = {
                        '目的地': f"{destination} ({end['name']})",
                        '類型': '國內-開車',
                        '距離(km)': distance_km,
                        '時間': f"{duration_min}分鐘",
                        '起點': start,
                        '終點': end,
                        'route': route['geometry']
                    }
                    
                    return {
                        '序號': row['序號'],
                        '傳票編號': row['傳票編號'],
                        '目的地': f"{destination} ({end['name']})",
                        '類型': '國內-開車',
                        '距離(km)': distance_km,
                        '時間': f"{duration_min}分鐘",
                        '起點': start,
                        '終點': end,
                        'route': route['geometry'],
                        'route_key': route_key,
                        'cached': False
                    }
        except:
            pass
            
        # 如果API失敗，使用直線距離估算
        start = COORDINATES['輔仁大學']
        end = COORDINATES[destination]
        start_point = (start['lat'], start['lon'])
        end_point = (end['lat'], end['lon'])
        distance_km = round(geodesic(start_point, end_point).kilometers * 1.4, 1)
        duration_min = round(distance_km * 1.5)
        
        # 儲存到快取
        self.route_cache[route_key] = {
            '目的地': f"{destination} ({end['name']})",
            '類型': '國內-開車',
            '距離(km)': distance_km,
            '時間': f"{duration_min}分鐘",
            '起點': start,
            '終點': end,
            'route': None
        }
        
        return {
            '序號': row['序號'],
            '傳票編號': row['傳票編號'],
            '目的地': f"{destination} ({end['name']})",
            '類型': '國內-開車',
            '距離(km)': distance_km,
            '時間': f"{duration_min}分鐘",
            '起點': start,
            '終點': end,
            'route': None,
            'route_key': route_key,
            'cached': False
        }
        
    def _calculate_island_flight_distance(self, row, destination):
        """計算台灣離島航班距離（從松山機場或桃園機場）"""
        if destination not in COORDINATES:
            return None
        
        # 根據離島選擇合適的起飛機場
        if destination in ['澎湖', '金門', '馬祖'] or (destination in COORDINATES and COORDINATES[destination]['type'] == 'island'):
            # 離島通常從松山機場起飛
            start = COORDINATES['松山機場']
            flight_type = '國內-離島航班'
        else:
            start = COORDINATES['桃園機場']
            flight_type = '國內-航班'
        
        # 建立路線的唯一識別碼
        route_key = f"{start['name']}-{destination}-island-flight"
        
        # 檢查快取
        if route_key in self.route_cache:
            cached_data = self.route_cache[route_key]
            return {
                '序號': row['序號'],
                '傳票編號': row['傳票編號'],
                '目的地': cached_data['目的地'],
                '類型': cached_data['類型'],
                '距離(km)': cached_data['距離(km)'],
                '時間': cached_data['時間'],
                '起點': cached_data['起點'],
                '終點': cached_data['終點'],
                'route': None,
                'route_key': route_key,
                'cached': True
            }
            
        # 計算離島航班距離
        end = COORDINATES[destination]
        
        # 計算大圓距離
        start_point = (start['lat'], start['lon'])
        end_point = (end['lat'], end['lon'])
        distance_km = round(geodesic(start_point, end_point).kilometers, 1)
        
        # 估算飛行時間（離島航班通常較短）
        flight_hours = distance_km / 500 + 0.5  # 離島航班速度較慢
        
        # 儲存到快取
        self.route_cache[route_key] = {
            '目的地': f"{destination} ({end['name']})",
            '類型': flight_type,
            '距離(km)': distance_km,
            '時間': f"{flight_hours:.1f}小時",
            '起點': start,
            '終點': end
        }
        
        return {
            '序號': row['序號'],
            '傳票編號': row['傳票編號'],
            '目的地': f"{destination} ({end['name']})",
            '類型': flight_type,
            '距離(km)': distance_km,
            '時間': f"{flight_hours:.1f}小時",
            '起點': start,
            '終點': end,
            'route': None,
            'route_key': route_key,
            'cached': False
        }
        
    def _calculate_flight_distance(self, row, destination):
        """計算國際飛行距離（統一從桃園機場出發）"""
        if destination not in COORDINATES:
            return None
        
        # 特殊處理港澳地區（直接使用當地機場）
        if destination in ['香港', '澳門']:
            start = COORDINATES['桃園機場']
            end = COORDINATES[destination]
            flight_type = '港澳-飛行'
        else:
            # 其他國外目的地統一從桃園機場出發
            start = COORDINATES['桃園機場']
            end = COORDINATES[destination]
            flight_type = '國際-飛行'
        
        # 建立路線的唯一識別碼
        route_key = f"桃園機場-{destination}-flying"
        
        # 檢查快取
        if route_key in self.route_cache:
            cached_data = self.route_cache[route_key]
            return {
                '序號': row['序號'],
                '傳票編號': row['傳票編號'],
                '目的地': cached_data['目的地'],
                '類型': cached_data['類型'],
                '距離(km)': cached_data['距離(km)'],
                '時間': cached_data['時間'],
                '起點': cached_data['起點'],
                '終點': cached_data['終點'],
                'route': None,
                'route_key': route_key,
                'cached': True
            }
        
        # 計算大圓距離
        start_point = (start['lat'], start['lon'])
        end_point = (end['lat'], end['lon'])
        distance_km = round(geodesic(start_point, end_point).kilometers, 1)
        
        # 估算飛行時間（包含起降）
        if destination in ['香港', '澳門']:
            flight_hours = distance_km / 700 + 0.3  # 港澳航班較短
        else:
            flight_hours = distance_km / 800 + 0.5  # 國際航班
        
        # 儲存到快取
        self.route_cache[route_key] = {
            '目的地': f"{destination} ({end['name']})",
            '類型': flight_type,
            '距離(km)': distance_km,
            '時間': f"{flight_hours:.1f}小時",
            '起點': start,
            '終點': end
        }
        
        return {
            '序號': row['序號'],
            '傳票編號': row['傳票編號'],
            '目的地': f"{destination} ({end['name']})",
            '類型': flight_type,
            '距離(km)': distance_km,
            '時間': f"{flight_hours:.1f}小時",
            '起點': start,
            '終點': end,
            'route': None,
            'route_key': route_key,
            'cached': False
        }
        
    def _add_result_to_tree(self, result):
        """添加結果到Treeview"""
        values = (
            result['序號'],
            result['傳票編號'],
            result['目的地'],
            result['類型'],
            result['距離(km)'],
            result['時間'],
            '待產生',
            '✓' if result.get('cached') else ''
        )
        self.tree.insert('', 'end', values=values)
        
    def generate_maps(self):
        """產生地圖"""
        if not self.results:
            messagebox.showerror("錯誤", "請先計算距離")
            return
            
        # 在新執行緒中執行
        thread = threading.Thread(target=self._generate_maps_thread)
        thread.start()
        
    def _generate_maps_thread(self):
        """產生地圖的執行緒函數"""
        try:
            # 確保輸出目錄存在
            map_dir = os.path.join(self.output_dir.get(), 'maps')
            os.makedirs(map_dir, exist_ok=True)
            
            total = len(self.results)
            unique_routes = {}  # 記錄唯一路線
            cache_hit = 0
            new_maps = 0
            
            # 先識別所有唯一路線
            for result in self.results:
                route_key = result.get('route_key', f"{result['起點']['name']}-{result['終點']['name']}")
                if route_key not in unique_routes:
                    unique_routes[route_key] = []
                unique_routes[route_key].append(result)
            
            self.status_text.set(f"發現 {len(unique_routes)} 條不同路線")
            
            # 產生地圖
            for idx, result in enumerate(self.results):
                self.status_text.set(f"正在處理 {idx+1}/{total}...")
                self.progress.set((idx + 1) / total * 100)
                
                route_key = result.get('route_key', f"{result['起點']['name']}-{result['終點']['name']}")
                
                # 檢查是否已有相同路線的地圖
                if route_key in self.map_cache and os.path.exists(self.map_cache[route_key]):
                    # 使用已存在的地圖檔案
                    result['地圖檔案'] = self.map_cache[route_key]
                    self.root.after(0, self._update_tree_item, idx, '已使用快取')
                    cache_hit += 1
                else:
                    # 產生新地圖
                    map_file = self._create_map(result, map_dir, f"{route_key}_{idx+1}")
                    
                    if map_file:
                        # 截圖
                        screenshot = self._capture_map_screenshot(map_file)
                        if screenshot:
                            result['地圖檔案'] = screenshot
                            # 儲存到快取
                            self.map_cache[route_key] = screenshot
                        else:
                            result['地圖檔案'] = map_file
                            self.map_cache[route_key] = map_file
                            
                        self.root.after(0, self._update_tree_item, idx, '已產生')
                        new_maps += 1
                        
            # 儲存快取
            self.save_cache()
            
            self.status_text.set(f"地圖產生完成！共 {len(unique_routes)} 條路線，快取 {cache_hit} 個，新建 {new_maps} 個")
            self.progress.set(100)
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("錯誤", f"產生地圖失敗: {str(e)}"))
            self.status_text.set("產生地圖失敗")
            
    def _create_map(self, result, output_dir, name):
        """創建地圖（包含智能縮放）"""
        try:
            # 取得起終點座標
            start_lat = result['起點']['lat']
            start_lon = result['起點']['lon']
            end_lat = result['終點']['lat']
            end_lon = result['終點']['lon']
            
            # 計算中心點
            center_lat = (start_lat + end_lat) / 2
            center_lon = (start_lon + end_lon) / 2
            
            # 根據距離決定縮放級別
            distance = result.get('距離(km)', 0)
            zoom_level = calculate_zoom_level(distance)
            
            # 創建地圖
            m = folium.Map(
                location=[center_lat, center_lon], 
                zoom_start=zoom_level,
                tiles='OpenStreetMap',
                prefer_canvas=True
            )
            
            # 添加起點標記（綠色）
            folium.Marker(
                [start_lat, start_lon],
                popup=f"<b>起點</b><br>{result['起點']['name']}<br>輔仁大學正門",
                icon=folium.Icon(color='green', icon='play', prefix='fa'),
                tooltip="起點"
            ).add_to(m)
            
            # 添加終點標記（紅色）
            folium.Marker(
                [end_lat, end_lon],
                popup=f"<b>終點</b><br>{result['終點']['name']}",
                icon=folium.Icon(color='red', icon='stop', prefix='fa'),
                tooltip="終點"
            ).add_to(m)
            
            # 添加路線
            if result['類型'] == '國內-開車' and result.get('route'):
                # 開車路線（詳細路徑）
                route_coords = [[coord[1], coord[0]] for coord in result['route']['coordinates']]
                
                # 主路線
                folium.PolyLine(
                    route_coords,
                    color='#0066CC',
                    weight=6,
                    opacity=0.8,
                    smooth_factor=1
                ).add_to(m)
                
                # 路線邊框（讓路線更明顯）
                folium.PolyLine(
                    route_coords,
                    color='#000033',
                    weight=8,
                    opacity=0.4,
                    smooth_factor=1
                ).add_to(m)
                
                # 使用路線邊界來調整地圖視野
                if len(route_coords) > 0:
                    lats = [coord[0] for coord in route_coords]
                    lons = [coord[1] for coord in route_coords]
                    sw = [min(lats), min(lons)]
                    ne = [max(lats), max(lons)]
                    
                    # 根據距離調整padding
                    if distance < 20:
                        padding = 0.02  # 短距離用較小的padding
                    elif distance < 50:
                        padding = 0.05
                    else:
                        padding = 0.1
                        
                    m.fit_bounds([sw, ne], padding=(padding, padding))
                    
            else:
                # 飛行路線（曲線）
                # 添加飛行路線的曲線效果（貝茲曲線）
                import numpy as np
                t = np.linspace(0, 1, 50)
                
                # 計算控制點（讓路線呈現弧形）
                control_lat = center_lat + (distance / 10000) * 5  # 根據距離調整弧度
                control_lon = center_lon
                
                # 二次貝茲曲線
                curve_lats = (1-t)**2 * start_lat + 2*(1-t)*t * control_lat + t**2 * end_lat
                curve_lons = (1-t)**2 * start_lon + 2*(1-t)*t * control_lon + t**2 * end_lon
                
                curve_coords = [[lat, lon] for lat, lon in zip(curve_lats, curve_lons)]
                
                # 飛行路線（橘色曲線）
                folium.PolyLine(
                    curve_coords,
                    color='#FF6600',
                    weight=4,
                    opacity=0.8
                ).add_to(m)
                
                # 虛線直線（參考用）
                folium.PolyLine(
                    [[start_lat, start_lon], [end_lat, end_lon]],
                    color='#CC0000',
                    weight=2,
                    opacity=0.4,
                    dash_array='10, 5'
                ).add_to(m)
                
                # 確保兩個機場都在視野內
                sw = [min(start_lat, end_lat), min(start_lon, end_lon)]
                ne = [max(start_lat, end_lat), max(start_lon, end_lon)]
                
                # 國際航線需要更大的padding
                padding = 0.2 if distance > 1000 else 0.1
                m.fit_bounds([sw, ne], padding=(padding, padding))
            
            # 添加距離標記（在路線中點）
            folium.Marker(
                [center_lat, center_lon],
                icon=folium.DivIcon(
                    html=f"""
                    <div style="background-color: white; 
                                border: 2px solid black; 
                                border-radius: 5px; 
                                padding: 5px;
                                font-weight: bold;
                                font-size: 12px;">
                        {distance} km
                    </div>
                    """
                )
            ).add_to(m)
            
            # 添加資訊框（固定在右上角）
            info_html = f"""
            <div style='position: fixed; 
                        top: 10px; right: 10px; 
                        background: rgba(255,255,255,0.95); 
                        padding: 15px;
                        border: 2px solid #333;
                        border-radius: 8px;
                        font-family: "Microsoft JhengHei", Arial;
                        box-shadow: 0 2px 6px rgba(0,0,0,0.3);
                        z-index: 9999;
                        max-width: 300px;'>
                <h3 style='margin: 0 0 10px 0; color: #333;'>路線資訊</h3>
                <table style='font-size: 14px;'>
                    <tr><td style='padding: 3px;'><b>傳票編號:</b></td><td>{result.get('傳票編號', 'N/A')}</td></tr>
                    <tr><td style='padding: 3px;'><b>起點:</b></td><td>{result['起點']['name']}</td></tr>
                    <tr><td style='padding: 3px;'><b>終點:</b></td><td>{result['終點']['name']}</td></tr>
                    <tr><td style='padding: 3px;'><b>距離:</b></td><td style='color: #0066CC; font-weight: bold;'>{distance} km</td></tr>
                    <tr><td style='padding: 3px;'><b>時間:</b></td><td>{result.get('時間', 'N/A')}</td></tr>
                    <tr><td style='padding: 3px;'><b>類型:</b></td><td>{result.get('類型', 'N/A')}</td></tr>
                </table>
            </div>
            """
            m.get_root().html.add_child(folium.Element(info_html))
            
            # 儲存地圖
            filename = f"map_{name}.html"
            filepath = os.path.join(output_dir, filename)
            m.save(filepath)
            
            return filepath
            
        except Exception as e:
            print(f"創建地圖失敗: {e}")
            return None
            
    def _capture_map_screenshot(self, html_file):
        """截取地圖截圖"""
        try:
            # 設置Chrome選項
            options = Options()
            options.add_argument('--headless')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--window-size=1200,800')
            
            # 創建瀏覽器實例
            driver = webdriver.Chrome(options=options)
            
            # 載入地圖
            driver.get(f"file://{html_file}")
            
            # 等待地圖載入
            time.sleep(3)
            
            # 截圖
            screenshot_path = html_file.replace('.html', '.png')
            driver.save_screenshot(screenshot_path)
            
            driver.quit()
            
            return screenshot_path
            
        except Exception as e:
            print(f"截圖失敗: {e}")
            # 如果Chrome截圖失敗，返回None
            return None
            
    def _update_tree_item(self, index, status):
        """更新Treeview項目"""
        items = self.tree.get_children()
        if index < len(items):
            item = items[index]
            values = list(self.tree.item(item)['values'])
            values[-2] = status  # 地圖狀態
            self.tree.item(item, values=values)
            
    def export_excel(self):
        """匯出Excel"""
        if not self.results:
            messagebox.showerror("錯誤", "沒有可匯出的資料")
            return
            
        try:
            self.status_text.set("正在匯出Excel...")
            
            # 建立DataFrame
            export_data = []
            for result in self.results:
                export_data.append({
                    '序號': result['序號'],
                    '傳票編號': result['傳票編號'],
                    '目的地': result['目的地'],
                    '類型': result['類型'],
                    '距離(km)': result['距離(km)'],
                    '時間': result['時間'],
                    '起點': result['起點']['name'],
                    '終點': result['終點']['name']
                })
                
            df = pd.DataFrame(export_data)
            
            # 儲存Excel
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(self.output_dir.get(), f'出差距離計算結果_{timestamp}.xlsx')
            
            # 使用ExcelWriter來插入圖片
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='距離計算結果', index=False)
                
                # 取得工作表
                workbook = writer.book
                worksheet = writer.sheets['距離計算結果']
                
                # 調整欄寬
                column_widths = {'A': 10, 'B': 15, 'C': 30, 'D': 15, 'E': 12, 'F': 12, 'G': 20, 'H': 20}
                for column, width in column_widths.items():
                    worksheet.column_dimensions[column].width = width
                    
                # 插入地圖截圖（如果有的話）
                if any('地圖檔案' in r for r in self.results):
                    # 新增一個工作表放地圖
                    ws_maps = workbook.create_sheet('地圖')
                    
                    # 按照唯一路線分組
                    route_map = {}
                    for result in self.results:
                        if '地圖檔案' not in result:
                            continue
                        route_key = result.get('route_key', f"{result['起點']['name']}-{result['終點']['name']}")
                        if route_key not in route_map:
                            route_map[route_key] = []
                        route_map[route_key].append(result)
                    
                    row = 1
                    for route_key, results in route_map.items():
                        # 只為每條唯一路線插入一次地圖
                        result = results[0]  # 取第一個結果
                        if '地圖檔案' in result and result['地圖檔案'] and os.path.exists(result['地圖檔案']):
                            # 添加標題
                            ws_maps.cell(row=row, column=1, value=f"路線: {result['起點']['name']} → {result['終點']['name']}")
                            ws_maps.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True, size=14)
                            row += 1
                            
                            # 列出使用此路線的所有傳票
                            vouchers = [str(r['傳票編號']) for r in results]  # 確保轉換為字串
                            ws_maps.cell(row=row, column=1, value=f"傳票編號: {', '.join(vouchers)}")
                            row += 1
                            
                            # 列出距離和時間
                            ws_maps.cell(row=row, column=1, value=f"距離: {result['距離(km)']} km | 時間: {result['時間']}")
                            row += 1
                            
                            # 插入圖片
                            try:
                                if result['地圖檔案'].endswith('.png'):
                                    img = XLImage(result['地圖檔案'])
                                    img.width = 600
                                    img.height = 400
                                    ws_maps.add_image(img, f'A{row}')
                                    row += 25  # 預留圖片空間
                                else:
                                    ws_maps.cell(row=row, column=1, value=f"地圖檔案: {result['地圖檔案']}")
                                    row += 2
                            except Exception as e:
                                print(f"插入圖片失敗: {e}")
                                ws_maps.cell(row=row, column=1, value=f"地圖檔案: {result['地圖檔案']}")
                                row += 2
                            
                            # 添加分隔線
                            row += 2
                            
            messagebox.showinfo("成功", f"Excel已匯出至:\n{output_file}")
            self.status_text.set("匯出完成")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"匯出失敗: {str(e)}")
            self.status_text.set("匯出失敗")
            
    def run(self):
        """執行GUI"""
        print("正在啟動GUI視窗...")
        self.root.update_idletasks()
        self.root.update()
        self.root.mainloop()
        print("GUI視窗已關閉")


# 主程式
if __name__ == "__main__":
    print("=== 出差距離計算與地圖產生工具 ===")
    print("檢查環境...")
    
    # 確保必要的套件已安裝
    required_packages = {
        'pandas': 'pandas',
        'folium': 'folium',
        'selenium': 'selenium',
        'PIL': 'pillow',
        'geopy': 'geopy',
        'openpyxl': 'openpyxl',
        'requests': 'requests'
    }
    
    missing_packages = []
    for import_name, install_name in required_packages.items():
        try:
            __import__(import_name)
        except ImportError:
            missing_packages.append(install_name)
            
    if missing_packages:
        print("\n錯誤：缺少必要的套件")
        print("請執行以下指令安裝:")
        print(f"pip install {' '.join(missing_packages)}")
        sys.exit(1)
    
    print("✓ 所有套件已就緒")
    print("\n正在啟動應用程式...")
    
    try:
        # 啟動應用程式
        app = TravelDistanceCalculator()
        print("✓ 應用程式已初始化")
        print("\n如果沒有看到視窗，請檢查：")
        print("1. 是否有其他視窗擋住")
        print("2. 在 macOS 上檢查 Dock 或 Mission Control")
        print("3. 嘗試 Alt+Tab (Windows) 或 Cmd+Tab (macOS) 切換視窗")
        print("\n按 Ctrl+C 可以終止程式")
        app.run()
    except KeyboardInterrupt:
        print("\n程式已被使用者中斷")
    except Exception as e:
        print(f"\n錯誤：{e}")
        import traceback
        traceback.print_exc()
