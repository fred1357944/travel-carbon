#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
出差距離計算與地圖產生工具 (GUI版本 - 效率優化版)
主要優化：
1. 先找出所有唯一路線（最大公因數概念）
2. 只為唯一路線產生地圖，避免重複處理
3. 改進的進度顯示
"""

import os
import re
import sys
import time
import json
import pickle
import requests
import pandas as pd
import numpy as np
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
from collections import defaultdict
try:
    from PIL import Image, ImageTk
except ImportError:
    Image = None
    ImageTk = None
import folium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from geopy.distance import geodesic
import warnings
warnings.filterwarnings('ignore')

# 座標資料庫
COORDINATES = {
    # 台灣地區（政府機關位置）
    '輔仁大學': {'lat': 25.0356, 'lon': 121.4320, 'name': '輔仁大學正門', 'type': 'start'},
    '台北': {'lat': 25.0330, 'lon': 121.5654, 'name': '台北市政府', 'type': 'domestic'},
    '臺北': {'lat': 25.0330, 'lon': 121.5654, 'name': '台北市政府', 'type': 'domestic'},
    '新北': {'lat': 25.0119, 'lon': 121.4653, 'name': '新北市政府', 'type': 'domestic'},
    '桃園': {'lat': 24.9936, 'lon': 121.3010, 'name': '桃園市政府', 'type': 'domestic'},
    '新竹': {'lat': 24.8138, 'lon': 120.9675, 'name': '新竹市政府', 'type': 'domestic'},
    '苗栗': {'lat': 24.5602, 'lon': 120.8214, 'name': '苗栗縣政府', 'type': 'domestic'},
    '台中': {'lat': 24.1639, 'lon': 120.6478, 'name': '台中市政府', 'type': 'domestic'},
    '臺中': {'lat': 24.1639, 'lon': 120.6478, 'name': '台中市政府', 'type': 'domestic'},
    '彰化': {'lat': 24.0757, 'lon': 120.5442, 'name': '彰化縣政府', 'type': 'domestic'},
    '南投': {'lat': 23.9611, 'lon': 120.9719, 'name': '南投縣政府', 'type': 'domestic'},
    '雲林': {'lat': 23.7092, 'lon': 120.5410, 'name': '雲林縣政府', 'type': 'domestic'},
    '嘉義': {'lat': 23.4801, 'lon': 120.4538, 'name': '嘉義市政府', 'type': 'domestic'},
    '台南': {'lat': 22.9997, 'lon': 120.2269, 'name': '台南市政府', 'type': 'domestic'},
    '臺南': {'lat': 22.9997, 'lon': 120.2269, 'name': '台南市政府', 'type': 'domestic'},
    '高雄': {'lat': 22.6273, 'lon': 120.3014, 'name': '高雄市政府', 'type': 'domestic'},
    '屏東': {'lat': 22.6821, 'lon': 120.4871, 'name': '屏東縣政府', 'type': 'domestic'},
    '宜蘭': {'lat': 24.7021, 'lon': 121.7377, 'name': '宜蘭縣政府', 'type': 'domestic'},
    '花蓮': {'lat': 23.9871, 'lon': 121.6015, 'name': '花蓮縣政府', 'type': 'domestic'},
    '台東': {'lat': 22.7972, 'lon': 121.0713, 'name': '台東縣政府', 'type': 'domestic'},
    '臺東': {'lat': 22.7972, 'lon': 121.0713, 'name': '台東縣政府', 'type': 'domestic'},
    '澎湖': {'lat': 23.5711, 'lon': 119.5793, 'name': '澎湖縣政府', 'type': 'island'},
    '金門': {'lat': 24.4493, 'lon': 118.3766, 'name': '金門縣政府', 'type': 'island'},
    '馬祖': {'lat': 26.1505, 'lon': 119.9498, 'name': '連江縣政府', 'type': 'island'},
    '基隆': {'lat': 25.1324, 'lon': 121.7391, 'name': '基隆市政府', 'type': 'domestic'},
    
    # 台北地區車站與地點
    '南港': {'lat': 25.0530, 'lon': 121.6070, 'name': '南港車站', 'type': 'domestic', 'parent': '台北'},
    '松山': {'lat': 25.0491, 'lon': 121.5776, 'name': '松山車站', 'type': 'domestic', 'parent': '台北'},
    '信義': {'lat': 25.0330, 'lon': 121.5680, 'name': '信義區', 'type': 'domestic', 'parent': '台北'},
    '大安': {'lat': 25.0261, 'lon': 121.5435, 'name': '大安區', 'type': 'domestic', 'parent': '台北'},
    '中山': {'lat': 25.0692, 'lon': 121.5443, 'name': '中山區', 'type': 'domestic', 'parent': '台北'},
    '內湖': {'lat': 25.0835, 'lon': 121.5948, 'name': '內湖區', 'type': 'domestic', 'parent': '台北'},
    '士林': {'lat': 25.0928, 'lon': 121.5194, 'name': '士林區', 'type': 'domestic', 'parent': '台北'},
    '北投': {'lat': 25.1320, 'lon': 121.5034, 'name': '北投區', 'type': 'domestic', 'parent': '台北'},
    '政大': {'lat': 24.9857, 'lon': 121.5756, 'name': '國立政治大學', 'type': 'domestic'},
    '台大醫院': {'lat': 25.0403, 'lon': 121.5188, 'name': '台大醫院', 'type': 'domestic'},
    '新光醫院': {'lat': 25.0940, 'lon': 121.5205, 'name': '新光吳火獅紀念醫院', 'type': 'domestic'},
    
    # 新北地區
    '板橋': {'lat': 25.0146, 'lon': 121.4639, 'name': '板橋車站', 'type': 'domestic', 'parent': '新北'},
    '三重': {'lat': 25.0703, 'lon': 121.4876, 'name': '三重區', 'type': 'domestic', 'parent': '新北'},
    '中和': {'lat': 25.0020, 'lon': 121.4995, 'name': '中和區', 'type': 'domestic', 'parent': '新北'},
    '永和': {'lat': 25.0170, 'lon': 121.5140, 'name': '永和區', 'type': 'domestic', 'parent': '新北'},
    '新店': {'lat': 24.9578, 'lon': 121.5375, 'name': '新店區', 'type': 'domestic', 'parent': '新北'},
    '汐止': {'lat': 25.0630, 'lon': 121.6620, 'name': '汐止區', 'type': 'domestic', 'parent': '新北'},
    '淡水': {'lat': 25.1681, 'lon': 121.4454, 'name': '淡水區', 'type': 'domestic', 'parent': '新北'},
    '五股': {'lat': 25.0856, 'lon': 121.4383, 'name': '五股區', 'type': 'domestic', 'parent': '新北'},
    '三峽': {'lat': 24.9342, 'lon': 121.3686, 'name': '三峽區', 'type': 'domestic', 'parent': '新北'},
    '淡江': {'lat': 25.1748, 'lon': 121.4507, 'name': '淡江大學', 'type': 'domestic'},
    '新莊': {'lat': 25.0360, 'lon': 121.4525, 'name': '新莊區', 'type': 'domestic', 'parent': '新北'},
    '鶯歌': {'lat': 24.9549, 'lon': 121.3547, 'name': '鶯歌區', 'type': 'domestic', 'parent': '新北'},
    '樹林': {'lat': 24.9913, 'lon': 121.4245, 'name': '樹林區', 'type': 'domestic', 'parent': '新北'},
    '深坑': {'lat': 24.9915, 'lon': 121.6157, 'name': '深坑區', 'type': 'domestic', 'parent': '新北'},
    '石碇': {'lat': 24.9914, 'lon': 121.6585, 'name': '石碇區', 'type': 'domestic', 'parent': '新北'},
    '坪林': {'lat': 24.9371, 'lon': 121.7119, 'name': '坪林區', 'type': 'domestic', 'parent': '新北'},
    '林口': {'lat': 25.0776, 'lon': 121.3917, 'name': '林口區', 'type': 'domestic', 'parent': '新北'},
    '土城': {'lat': 24.9725, 'lon': 121.4431, 'name': '土城區', 'type': 'domestic', 'parent': '新北'},
    '蘆洲': {'lat': 25.0849, 'lon': 121.4734, 'name': '蘆洲區', 'type': 'domestic', 'parent': '新北'},
    '貢寮': {'lat': 25.0217, 'lon': 121.9089, 'name': '貢寮區', 'type': 'domestic', 'parent': '新北'},
    '瑞芳': {'lat': 25.1089, 'lon': 121.8107, 'name': '瑞芳區', 'type': 'domestic', 'parent': '新北'},
    '雙溪': {'lat': 25.0340, 'lon': 121.8657, 'name': '雙溪區', 'type': 'domestic', 'parent': '新北'},
    '平溪': {'lat': 25.0259, 'lon': 121.7584, 'name': '平溪區', 'type': 'domestic', 'parent': '新北'},
    '烏來': {'lat': 24.8651, 'lon': 121.5506, 'name': '烏來區', 'type': 'domestic', 'parent': '新北'},
    '金山': {'lat': 25.2219, 'lon': 121.6367, 'name': '金山區', 'type': 'domestic', 'parent': '新北'},
    '萬里': {'lat': 25.1766, 'lon': 121.6887, 'name': '萬里區', 'type': 'domestic', 'parent': '新北'},
    '八里': {'lat': 25.1467, 'lon': 121.3982, 'name': '八里區', 'type': 'domestic', 'parent': '新北'},
    '泰山': {'lat': 25.0589, 'lon': 121.4306, 'name': '泰山區', 'type': 'domestic', 'parent': '新北'},
    '鹽宮': {'lat': 24.9685, 'lon': 121.7085, 'name': '鹽宮', 'type': 'domestic', 'parent': '新北'},
    
    # 其他大學
    '東海大學': {'lat': 24.1795, 'lon': 120.6003, 'name': '東海大學', 'type': 'domestic'},
    '玄奘大學': {'lat': 24.8314, 'lon': 120.9444, 'name': '玄奘大學', 'type': 'domestic'},
    '大同大學': {'lat': 25.0648, 'lon': 121.5162, 'name': '大同大學', 'type': 'domestic'},
    '佛光大學': {'lat': 24.8597, 'lon': 121.7225, 'name': '佛光大學', 'type': 'domestic'},
    '布達佩斯': {'lat': 47.4298, 'lon': 19.2611, 'name': '布達佩斯李斯特·費倫茨國際機場', 'type': 'international'},
    '關渡': {'lat': 25.1247, 'lon': 121.4674, 'name': '關渡', 'type': 'domestic', 'parent': '台北'},
    '中山大學': {'lat': 22.6288, 'lon': 120.2649, 'name': '國立中山大學', 'type': 'domestic'},
    '虎尾科大': {'lat': 23.7012, 'lon': 120.5351, 'name': '國立虎尾科技大學', 'type': 'domestic'},
    '靜宜大學': {'lat': 24.2451, 'lon': 120.5817, 'name': '靜宜大學', 'type': 'domestic'},
    '中原大學': {'lat': 24.9572, 'lon': 121.2407, 'name': '中原大學', 'type': 'domestic'},
    '逢甲': {'lat': 24.1785, 'lon': 120.6465, 'name': '逢甲大學', 'type': 'domestic'},
    '靜宜': {'lat': 24.2451, 'lon': 120.5817, 'name': '靜宜大學', 'type': 'domestic'},
    '清華大學': {'lat': 24.7958, 'lon': 120.9950, 'name': '國立清華大學', 'type': 'domestic'},
    '陽明交大': {'lat': 24.7868, 'lon': 120.9975, 'name': '陽明交通大學', 'type': 'domestic'},
    '明道': {'lat': 23.8679, 'lon': 120.4708, 'name': '明道大學', 'type': 'domestic'},
    '中山醫': {'lat': 24.1232, 'lon': 120.6508, 'name': '中山醫學大學', 'type': 'domestic'},
    '北醫': {'lat': 25.0269, 'lon': 121.5616, 'name': '臺北醫學大學', 'type': 'domestic'},
    '高醫': {'lat': 22.6450, 'lon': 120.3236, 'name': '高雄醫學大學', 'type': 'domestic'},
    '文化': {'lat': 25.1367, 'lon': 121.5304, 'name': '中國文化大學', 'type': 'domestic'},
    '東吳': {'lat': 25.1283, 'lon': 121.5118, 'name': '東吳大學', 'type': 'domestic'},
    '銘傳': {'lat': 25.0836, 'lon': 121.5251, 'name': '銘傳大學', 'type': 'domestic'},
    '耕莘護': {'lat': 25.0356, 'lon': 121.5227, 'name': '耕莘健康管理專科學校', 'type': 'domestic'},
    '南科大': {'lat': 23.0163, 'lon': 120.2238, 'name': '南臺科技大學', 'type': 'domestic'},
    '成功': {'lat': 22.9968, 'lon': 120.2219, 'name': '國立成功大學', 'type': 'domestic'},
    '元智': {'lat': 24.9701, 'lon': 121.2645, 'name': '元智大學', 'type': 'domestic'},
    '長庚': {'lat': 25.0825, 'lon': 121.3918, 'name': '長庚大學', 'type': 'domestic'},
    '中興': {'lat': 24.1217, 'lon': 120.6756, 'name': '國立中興大學', 'type': 'domestic'},
    '建國科': {'lat': 24.0707, 'lon': 120.5485, 'name': '建國科技大學', 'type': 'domestic'},
    '正修科': {'lat': 22.7323, 'lon': 120.2980, 'name': '正修科技大學', 'type': 'domestic'},
    '彰化師': {'lat': 24.0810, 'lon': 120.5582, 'name': '國立彰化師範大學', 'type': 'domestic'},
    '龍華科': {'lat': 25.0018, 'lon': 121.4168, 'name': '龍華科技大學', 'type': 'domestic'},
    '實踐': {'lat': 25.0783, 'lon': 121.5440, 'name': '實踐大學', 'type': 'domestic'},
    '台師': {'lat': 25.0265, 'lon': 121.5279, 'name': '國立臺灣師範大學', 'type': 'domestic'},
    '臺師': {'lat': 25.0265, 'lon': 121.5279, 'name': '國立臺灣師範大學', 'type': 'domestic'},
    '臺藝': {'lat': 25.0070, 'lon': 121.4476, 'name': '國立臺灣藝術大學', 'type': 'domestic'},
    '海洋': {'lat': 25.1504, 'lon': 121.7750, 'name': '國立臺灣海洋大學', 'type': 'domestic'},
    '元智大學': {'lat': 24.9701, 'lon': 121.2645, 'name': '元智大學', 'type': 'domestic'},
    '萬能科大': {'lat': 24.9960, 'lon': 121.2480, 'name': '萬能科技大學', 'type': 'domestic'},
    '明新科大': {'lat': 24.8655, 'lon': 121.0006, 'name': '明新科技大學', 'type': 'domestic'},
    '景文科大': {'lat': 24.9425, 'lon': 121.6282, 'name': '景文科技大學', 'type': 'domestic'},
    '國體大': {'lat': 25.0336, 'lon': 121.3890, 'name': '國立體育大學', 'type': 'domestic'},
    '世新大學': {'lat': 24.9597, 'lon': 121.5435, 'name': '世新大學', 'type': 'domestic'},
    '南台科大': {'lat': 23.0163, 'lon': 120.2238, 'name': '南臺科技大學', 'type': 'domestic'},
    '中興大學': {'lat': 24.1217, 'lon': 120.6756, 'name': '國立中興大學', 'type': 'domestic'},
    '朝陽大學': {'lat': 24.2438, 'lon': 120.6572, 'name': '朝陽科技大學', 'type': 'domestic'},
    '樹德科大': {'lat': 22.7323, 'lon': 120.3980, 'name': '樹德科技大學', 'type': 'domestic'},
    '高師大': {'lat': 22.6253, 'lon': 120.2827, 'name': '國立高雄師範大學', 'type': 'domestic'},
    '慈濟科大': {'lat': 23.9580, 'lon': 121.5496, 'name': '慈濟科技大學', 'type': 'domestic'},
    '輔仁大學': {'lat': 25.0356, 'lon': 121.4320, 'name': '輔仁大學', 'type': 'domestic'},
    '輔大醫院': {'lat': 25.0320, 'lon': 121.4377, 'name': '輔大醫院', 'type': 'domestic'},
    '臺大': {'lat': 25.0173, 'lon': 121.5397, 'name': '國立臺灣大學', 'type': 'domestic'},
    '衛理女中': {'lat': 25.0962, 'lon': 121.5498, 'name': '衛理女子高級中學', 'type': 'domestic'},
    '光仁高中': {'lat': 25.0103, 'lon': 121.4737, 'name': '光仁高級中學', 'type': 'domestic'},
    '陽明高中': {'lat': 25.1367, 'lon': 121.5338, 'name': '陽明高級中學', 'type': 'domestic'},
    '永平高中': {'lat': 24.9913, 'lon': 121.5126, 'name': '新北市立永平高級中學', 'type': 'domestic'},
    '泰山高中': {'lat': 25.0529, 'lon': 121.4306, 'name': '新北市立泰山高級中學', 'type': 'domestic'},
    '桃園高中': {'lat': 24.9936, 'lon': 121.3130, 'name': '國立桃園高級中學', 'type': 'domestic'},
    '新店高中': {'lat': 24.9723, 'lon': 121.5395, 'name': '新北市立新店高級中學', 'type': 'domestic'},
    '桃園振聲高中': {'lat': 24.9890, 'lon': 121.3150, 'name': '振聲高級中學', 'type': 'domestic'},
    '基隆高中': {'lat': 25.1256, 'lon': 121.7421, 'name': '國立基隆高級中學', 'type': 'domestic'},
    
    # 機場
    '桃園機場': {'lat': 25.0777, 'lon': 121.2325, 'name': '桃園國際機場', 'type': 'airport'},
    '松山機場': {'lat': 25.0694, 'lon': 121.5526, 'name': '台北松山機場', 'type': 'airport'},
    '澎湖機場': {'lat': 23.5676, 'lon': 119.6282, 'name': '澎湖機場', 'type': 'airport'},
    '金門機場': {'lat': 24.4274, 'lon': 118.3592, 'name': '金門機場', 'type': 'airport'},
    
    # 港澳地區
    '香港': {'lat': 22.3080, 'lon': 113.9185, 'name': '香港國際機場', 'type': 'international'},
    '澳門': {'lat': 22.1496, 'lon': 113.5910, 'name': '澳門國際機場', 'type': 'international'},
    
    # 新增特定地點（根據用戶建議）
    '泗水': {'lat': -7.3797, 'lon': 112.7913, 'name': '朱安達國際機場', 'type': 'international'},
    '浙江': {'lat': 30.2294, 'lon': 120.4344, 'name': '杭州蕭山國際機場', 'type': 'international'},
    '威尼斯': {'lat': 41.8002, 'lon': 12.2389, 'name': '羅馬菲烏米奇諾機場', 'type': 'international'},
    '羅馬': {'lat': 41.8002, 'lon': 12.2389, 'name': '羅馬菲烏米奇諾機場', 'type': 'international'},
    '山東': {'lat': 36.8577, 'lon': 117.2157, 'name': '濟南遙墻國際機場', 'type': 'international'},
    '費城': {'lat': 39.8719, 'lon': -75.2411, 'name': '費城國際機場', 'type': 'international'},
    '維也納': {'lat': 48.1103, 'lon': 16.5697, 'name': '維也納國際機場', 'type': 'international'},
    '布拉格': {'lat': 50.1008, 'lon': 14.2632, 'name': '布拉格瓦茨拉夫·哈維爾國際機場', 'type': 'international'},
    '蒙特婁': {'lat': 45.4706, 'lon': -73.7408, 'name': '蒙特婁皮埃爾·埃利奧特·特魯多國際機場', 'type': 'international'},
    '里司本': {'lat': -31.9402, 'lon': 115.9671, 'name': '珀斯機場', 'type': 'international'},
    
    # 特殊地點處理
    '東海': {'lat': 24.1795, 'lon': 120.6003, 'name': '東海大學', 'type': 'domestic'},
    '清華': {'lat': 24.7958, 'lon': 120.9950, 'name': '國立清華大學', 'type': 'domestic'},
    
    # 醫院
    '台中榮民總醫院': {'lat': 24.1886, 'lon': 120.6103, 'name': '臺中榮民總醫院', 'type': 'domestic'},
    '羅東聖母醫院': {'lat': 24.6901, 'lon': 121.7697, 'name': '羅東聖母醫院', 'type': 'domestic'},
    '新店慈濟醫院': {'lat': 24.9684, 'lon': 121.5373, 'name': '新店慈濟醫院', 'type': 'domestic'},
    '聖馬爾定醫院': {'lat': 23.4650, 'lon': 120.4558, 'name': '嘉義市聖馬爾定醫院', 'type': 'domestic'},
    '高雄榮總': {'lat': 22.6792, 'lon': 120.3233, 'name': '高雄榮民總醫院', 'type': 'domestic'},
    '高醫大': {'lat': 22.6450, 'lon': 120.3236, 'name': '高雄醫學大學附設中和紀念醫院', 'type': 'domestic'},
    '三總': {'lat': 25.0707, 'lon': 121.5397, 'name': '三軍總醫院', 'type': 'domestic'},
    '長庚': {'lat': 25.0055, 'lon': 121.3010, 'name': '長庚紀念醫院', 'type': 'domestic'},
    '亞東醫院': {'lat': 24.9760, 'lon': 121.4560, 'name': '亞東紀念醫院', 'type': 'domestic'},
    '馬偕專科': {'lat': 25.0750, 'lon': 121.3750, 'name': '馬偕醫護管理專科學校', 'type': 'domestic'},
    
    # 政府機關
    '教育部': {'lat': 25.0285, 'lon': 121.5205, 'name': '教育部', 'type': 'domestic'},
    '國科會': {'lat': 25.0143, 'lon': 121.5363, 'name': '國家科學及技術委員會', 'type': 'domestic'},
    '高等法院': {'lat': 25.0343, 'lon': 121.5061, 'name': '臺灣高等法院', 'type': 'domestic'},
    '僑委會': {'lat': 25.0406, 'lon': 121.5184, 'name': '僑務委員會', 'type': 'domestic'},
    '台北市政府': {'lat': 25.0330, 'lon': 121.5654, 'name': '台北市政府', 'type': 'domestic'},
    
    # 渡假村/活動中心
    '大板根': {'lat': 24.9000, 'lon': 121.4000, 'name': '大板根森林溫泉渡假村', 'type': 'domestic'},
    '劍潭青年活動中心': {'lat': 25.0792, 'lon': 121.5245, 'name': '劍潭青年活動中心', 'type': 'domestic'},
    '坪林合歡山莊': {'lat': 24.9380, 'lon': 121.7119, 'name': '坪林合歡山莊', 'type': 'domestic'},
    '龍門露營區': {'lat': 25.0005, 'lon': 121.9428, 'name': '龍門露營區', 'type': 'domestic'},
    '文山農場': {'lat': 24.8426, 'lon': 121.5826, 'name': '文山農場', 'type': 'domestic'},
    '宜蘭仁愛之家': {'lat': 24.7471, 'lon': 121.7460, 'name': '宜蘭仁愛之家', 'type': 'domestic'},
    '新莊水源地': {'lat': 25.0530, 'lon': 121.4315, 'name': '新莊水源地', 'type': 'domestic'},
    '大同育幼院': {'lat': 25.0734, 'lon': 121.5151, 'name': '大同育幼院', 'type': 'domestic'},
    
    # 科技公司
    '安圖斯科技': {'lat': 24.9792, 'lon': 121.5448, 'name': '安圖斯科技', 'type': 'domestic'},
    
    # 國外大學
    '上海體育大學': {'lat': 31.0761, 'lon': 121.2034, 'name': '上海體育大學', 'type': 'international'},
    
    # 全名別名
    '成功大學': {'lat': 22.9968, 'lon': 120.2219, 'name': '國立成功大學', 'type': 'domestic'},
    '靜宜大學': {'lat': 24.2451, 'lon': 120.5817, 'name': '靜宜大學', 'type': 'domestic'},
    '台科大': {'lat': 25.0085, 'lon': 121.5412, 'name': '國立臺灣科技大學', 'type': 'domestic'},
    '淡江大學': {'lat': 25.1748, 'lon': 121.4507, 'name': '淡江大學', 'type': 'domestic'},
    '文化大學': {'lat': 25.1367, 'lon': 121.5304, 'name': '中國文化大學', 'type': 'domestic'},
    '政治大學': {'lat': 24.9857, 'lon': 121.5756, 'name': '國立政治大學', 'type': 'domestic'},
    '東吳大學': {'lat': 25.1283, 'lon': 121.5118, 'name': '東吳大學', 'type': 'domestic'},
    
    # 大學別名
    '輔大': {'lat': 25.0356, 'lon': 121.4320, 'name': '輔仁大學', 'type': 'domestic'},
    '輔大醫院': {'lat': 25.0320, 'lon': 121.4377, 'name': '輔大醫院', 'type': 'domestic'},
    '臺大醫院': {'lat': 25.0403, 'lon': 121.5188, 'name': '臺大醫院', 'type': 'domestic'},
    '成大': {'lat': 22.9968, 'lon': 120.2219, 'name': '國立成功大學', 'type': 'domestic'},
    '高醫大': {'lat': 22.6450, 'lon': 120.3236, 'name': '高雄醫學大學', 'type': 'domestic'},
    '嘉義大學': {'lat': 23.5589, 'lon': 120.4723, 'name': '國立嘉義大學', 'type': 'domestic'},
    '嘉大': {'lat': 23.5589, 'lon': 120.4723, 'name': '國立嘉義大學', 'type': 'domestic'},
    
    # 其他地區地點
    '羅東': {'lat': 24.6818, 'lon': 121.7664, 'name': '羅東鎮', 'type': 'domestic'},
    '礁溪': {'lat': 24.8268, 'lon': 121.7736, 'name': '礁溪鄉', 'type': 'domestic'},
    '頭城': {'lat': 24.8593, 'lon': 121.8230, 'name': '頭城鎮', 'type': 'domestic'},
    '梅花湖': {'lat': 24.6412, 'lon': 121.7355, 'name': '梅花湖', 'type': 'domestic'},
    '大溪': {'lat': 24.8802, 'lon': 121.2869, 'name': '大溪區', 'type': 'domestic'},
    '沙鹿': {'lat': 24.2377, 'lon': 120.5579, 'name': '沙鹿區', 'type': 'domestic'},
    '新屋': {'lat': 24.9720, 'lon': 121.1058, 'name': '新屋區', 'type': 'domestic'},
    '大直': {'lat': 25.0795, 'lon': 121.5473, 'name': '大直', 'type': 'domestic'},
    '陽明山': {'lat': 25.1505, 'lon': 121.5363, 'name': '陽明山', 'type': 'domestic'},
    '鹽寮': {'lat': 24.8900, 'lon': 121.8357, 'name': '鹽寮', 'type': 'domestic'},
    '三富農場': {'lat': 24.9445, 'lon': 121.6542, 'name': '三富農場', 'type': 'domestic'},
    
    # 國際主要機場
    '日本': {'lat': 35.5494, 'lon': 139.7798, 'name': '成田國際機場', 'type': 'international'},
    '韓國': {'lat': 37.4602, 'lon': 126.4407, 'name': '仁川國際機場', 'type': 'international'},
    '新加坡': {'lat': 1.3644, 'lon': 103.9915, 'name': '樟宜機場', 'type': 'international'},
    '泰國': {'lat': 13.6900, 'lon': 100.7501, 'name': '素萬那普機場', 'type': 'international'},
    '美國': {'lat': 33.9425, 'lon': -118.4081, 'name': '洛杉磯國際機場', 'type': 'international'},
    '德國': {'lat': 50.0379, 'lon': 8.5622, 'name': '法蘭克福機場', 'type': 'international'},
    '英國': {'lat': 51.4700, 'lon': -0.4543, 'name': '希斯洛機場', 'type': 'international'},
    '法國': {'lat': 49.0097, 'lon': 2.5479, 'name': '戴高樂機場', 'type': 'international'},
    '澳洲': {'lat': -33.9399, 'lon': 151.1753, 'name': '雪梨機場', 'type': 'international'},
    '加拿大': {'lat': 43.6777, 'lon': -79.6248, 'name': '多倫多機場', 'type': 'international'},
    '印度': {'lat': 28.5562, 'lon': 77.1000, 'name': '英迪拉甘地機場', 'type': 'international'},
    '巴西': {'lat': -23.4356, 'lon': -46.4731, 'name': '聖保羅國際機場', 'type': 'international'},
    '墨西哥': {'lat': 19.4363, 'lon': -99.0721, 'name': '墨西哥城機場', 'type': 'international'},
    '印尼': {'lat': -6.1256, 'lon': 106.6559, 'name': '蘇加諾哈達機場', 'type': 'international'},
    '菲律賓': {'lat': 14.5086, 'lon': 121.0198, 'name': '尼諾伊艾奎諾機場', 'type': 'international'},
    '馬來西亞': {'lat': 2.7456, 'lon': 101.7072, 'name': '吉隆坡國際機場', 'type': 'international'},
    '越南': {'lat': 10.8184, 'lon': 106.6519, 'name': '新山一國際機場', 'type': 'international'},
    '迦納': {'lat': 5.6052, 'lon': -0.1668, 'name': '科托卡國際機場', 'type': 'international'},
    '南非': {'lat': -26.1367, 'lon': 28.2460, 'name': '約翰尼斯堡機場', 'type': 'international'},
    '埃及': {'lat': 30.1219, 'lon': 31.4056, 'name': '開羅國際機場', 'type': 'international'},
    '肯亞': {'lat': -1.3192, 'lon': 36.9278, 'name': '喬莫肯雅塔機場', 'type': 'international'},
    '衣索比亞': {'lat': 8.9779, 'lon': 38.7993, 'name': '博萊國際機場', 'type': 'international'},
    '紐西蘭': {'lat': -37.0082, 'lon': 174.7850, 'name': '奧克蘭國際機場', 'type': 'international'},
    '阿根廷': {'lat': -34.8222, 'lon': -58.5358, 'name': '埃塞薩國際機場', 'type': 'international'},
    '智利': {'lat': -33.3930, 'lon': -70.7858, 'name': '聖地牙哥國際機場', 'type': 'international'},
    '俄羅斯': {'lat': 55.9726, 'lon': 37.4146, 'name': '謝列梅捷沃國際機場', 'type': 'international'},
    '土耳其': {'lat': 40.9769, 'lon': 28.8146, 'name': '伊斯坦堡機場', 'type': 'international'},
    '阿聯酋': {'lat': 25.2532, 'lon': 55.3657, 'name': '杜拜國際機場', 'type': 'international'},
    '以色列': {'lat': 32.0114, 'lon': 34.8867, 'name': '本古里安機場', 'type': 'international'},
    '挪威': {'lat': 60.1939, 'lon': 11.1004, 'name': '奧斯陸機場', 'type': 'international'},
    '瑞士': {'lat': 47.4647, 'lon': 8.5492, 'name': '蘇黎世機場', 'type': 'international'},
    '瑞典': {'lat': 59.6519, 'lon': 17.9186, 'name': '斯德哥爾摩阿蘭達機場', 'type': 'international'},
    '丹麥': {'lat': 55.6179, 'lon': 12.6560, 'name': '哥本哈根機場', 'type': 'international'},
    '芬蘭': {'lat': 60.3172, 'lon': 24.9633, 'name': '赫爾辛基機場', 'type': 'international'},
    '波蘭': {'lat': 52.1657, 'lon': 20.9671, 'name': '華沙蕭邦機場', 'type': 'international'},
    '匈牙利': {'lat': 47.4298, 'lon': 19.2611, 'name': '布達佩斯機場', 'type': 'international'},
    '希臘': {'lat': 37.9364, 'lon': 23.9475, 'name': '雅典國際機場', 'type': 'international'},
    '比利時': {'lat': 50.9014, 'lon': 4.4844, 'name': '布魯塞爾機場', 'type': 'international'},
    '荷蘭': {'lat': 52.3086, 'lon': 4.7639, 'name': '阿姆斯特丹史基浦機場', 'type': 'international'},
    '葡萄牙': {'lat': 38.7813, 'lon': -9.1359, 'name': '里斯本機場', 'type': 'international'},
    '西班牙': {'lat': 40.4719, 'lon': -3.5626, 'name': '馬德里巴拉哈斯機場', 'type': 'international'},
    '義大利': {'lat': 41.8003, 'lon': 12.2389, 'name': '羅馬菲烏米奇諾機場', 'type': 'international'},
    '奧地利': {'lat': 48.1103, 'lon': 16.5697, 'name': '維也納國際機場', 'type': 'international'},
    '捷克': {'lat': 50.1008, 'lon': 14.2600, 'name': '布拉格瓦茨拉夫·哈維爾機場', 'type': 'international'},
    '羅馬尼亞': {'lat': 44.5711, 'lon': 26.0858, 'name': '布加勒斯特機場', 'type': 'international'},
    '斯洛維尼亞': {'lat': 46.2237, 'lon': 14.4576, 'name': '盧布爾雅那機場', 'type': 'international'},
    '塞爾維亞': {'lat': 44.8184, 'lon': 20.3091, 'name': '貝爾格勒機場', 'type': 'international'},
    '帛琉': {'lat': 7.3517, 'lon': 134.5444, 'name': '帛琉國際機場', 'type': 'international'},
    '坦尚尼亞': {'lat': -6.8781, 'lon': 39.2026, 'name': '朱利葉斯尼雷爾國際機場', 'type': 'international'},
    
    # 美國主要城市
    '洛杉磯': {'lat': 33.9425, 'lon': -118.4081, 'name': '洛杉磯國際機場', 'type': 'international'},
    '夏威夷': {'lat': 21.3186, 'lon': -157.9225, 'name': '檀香山國際機場', 'type': 'international'},
    '苝加哥': {'lat': 41.9742, 'lon': -87.9073, 'name': '歐黑爾國際機場', 'type': 'international'},
    '芝加哥': {'lat': 41.9742, 'lon': -87.9073, 'name': '歐黑爾國際機場', 'type': 'international'},
    '華盛頓': {'lat': 38.9531, 'lon': -77.4565, 'name': '杜勒斯國際機場', 'type': 'international'},
    '波士頓': {'lat': 42.3656, 'lon': -71.0096, 'name': '羅根國際機場', 'type': 'international'},
    '舊金山': {'lat': 37.6213, 'lon': -122.3790, 'name': '舊金山國際機場', 'type': 'international'},
    '西雅圖': {'lat': 47.4502, 'lon': -122.3088, 'name': '西雅圖塔科馬國際機場', 'type': 'international'},
    '鳳凰城': {'lat': 33.4373, 'lon': -112.0078, 'name': '鳳凰城天港國際機場', 'type': 'international'},
    '丹佛': {'lat': 39.8561, 'lon': -104.6737, 'name': '丹佛國際機場', 'type': 'international'},
    '印第安納州': {'lat': 39.7175, 'lon': -86.2950, 'name': '印第安納波利斯國際機場', 'type': 'international'},
    '波特蘭': {'lat': 45.5898, 'lon': -122.5951, 'name': '波特蘭國際機場', 'type': 'international'},
    
    # 中國主要機場
    '中國': {'lat': 40.0801, 'lon': 116.5846, 'name': '北京首都國際機場', 'type': 'international'},
    '北京': {'lat': 40.0801, 'lon': 116.5846, 'name': '北京首都國際機場', 'type': 'international'},
    '上海': {'lat': 31.1443, 'lon': 121.8083, 'name': '上海浦東國際機場', 'type': 'international'},
    '廣州': {'lat': 23.3924, 'lon': 113.2988, 'name': '廣州白雲國際機場', 'type': 'international'},
    '深圳': {'lat': 22.6393, 'lon': 113.8107, 'name': '深圳寶安國際機場', 'type': 'international'},
    '成都': {'lat': 30.5702, 'lon': 103.9471, 'name': '成都雙流國際機場', 'type': 'international'},
    '重慶': {'lat': 29.7196, 'lon': 106.6419, 'name': '重慶江北國際機場', 'type': 'international'},
    '杭州': {'lat': 30.2295, 'lon': 120.4346, 'name': '杭州蕭山國際機場', 'type': 'international'},
    '南京': {'lat': 31.7420, 'lon': 118.8620, 'name': '南京祿口國際機場', 'type': 'international'},
    '武漢': {'lat': 30.7839, 'lon': 114.2080, 'name': '武漢天河國際機場', 'type': 'international'},
    '西安': {'lat': 34.4471, 'lon': 108.7517, 'name': '西安咸陽國際機場', 'type': 'international'},
    '廈門': {'lat': 24.5440, 'lon': 118.1277, 'name': '廈門高崎國際機場', 'type': 'international'},
    '青島': {'lat': 36.2661, 'lon': 120.3743, 'name': '青島膠東國際機場', 'type': 'international'},
    '大連': {'lat': 38.9657, 'lon': 121.5386, 'name': '大連周水子國際機場', 'type': 'international'},
    '瀋陽': {'lat': 41.6398, 'lon': 123.4833, 'name': '瀋陽桃仙國際機場', 'type': 'international'},
    '哈爾濱': {'lat': 45.6236, 'lon': 126.2502, 'name': '哈爾濱太平國際機場', 'type': 'international'},
    '長沙': {'lat': 28.1892, 'lon': 113.2199, 'name': '長沙黃花國際機場', 'type': 'international'},
    '鄭州': {'lat': 34.5196, 'lon': 113.8408, 'name': '鄭州新鄭國際機場', 'type': 'international'},
    '昆明': {'lat': 25.0957, 'lon': 102.9285, 'name': '昆明長水國際機場', 'type': 'international'},
    '貴陽': {'lat': 26.5382, 'lon': 106.8004, 'name': '貴陽龍洞堡國際機場', 'type': 'international'},
    '南寧': {'lat': 22.6088, 'lon': 108.1727, 'name': '南寧吳圩國際機場', 'type': 'international'},
    '海口': {'lat': 19.9349, 'lon': 110.4589, 'name': '海口美蘭國際機場', 'type': 'international'},
    '三亞': {'lat': 18.3029, 'lon': 109.4122, 'name': '三亞鳳凰國際機場', 'type': 'international'},
    '蘭州': {'lat': 36.5152, 'lon': 103.6204, 'name': '蘭州中川國際機場', 'type': 'international'},
    '烏魯木齊': {'lat': 43.9071, 'lon': 87.4742, 'name': '烏魯木齊地窩堡國際機場', 'type': 'international'},
    '拉薩': {'lat': 29.2978, 'lon': 90.9119, 'name': '拉薩貢嘎國際機場', 'type': 'international'},
    '呼和浩特': {'lat': 40.8515, 'lon': 111.8241, 'name': '呼和浩特白塔國際機場', 'type': 'international'},
    '南昌': {'lat': 28.8650, 'lon': 115.9003, 'name': '南昌昌北國際機場', 'type': 'international'},
    '合肥': {'lat': 31.7800, 'lon': 117.2981, 'name': '合肥新橋國際機場', 'type': 'international'},
    '濟南': {'lat': 36.8572, 'lon': 117.2161, 'name': '濟南遙牆國際機場', 'type': 'international'},
    '福州': {'lat': 25.9351, 'lon': 119.6633, 'name': '福州長樂國際機場', 'type': 'international'},
    '太原': {'lat': 37.7469, 'lon': 112.6283, 'name': '太原武宿國際機場', 'type': 'international'},
    '石家莊': {'lat': 38.2807, 'lon': 114.6973, 'name': '石家莊正定國際機場', 'type': 'international'},
    '天津': {'lat': 39.1248, 'lon': 117.3462, 'name': '天津濱海國際機場', 'type': 'international'},
    
    # 印度主要城市
    '邦加羅爾': {'lat': 13.1986, 'lon': 77.7066, 'name': '肯佩戈達國際機場', 'type': 'international'},
    '班加羅爾': {'lat': 13.1986, 'lon': 77.7066, 'name': '肯佩戈達國際機場', 'type': 'international'},
    '德里': {'lat': 28.5562, 'lon': 77.1000, 'name': '英迪拉甘地機場', 'type': 'international'},
    '孟買': {'lat': 19.0896, 'lon': 72.8656, 'name': '賈特拉帕蒂希瓦吉國際機場', 'type': 'international'},
    '清奈': {'lat': 12.9941, 'lon': 80.1709, 'name': '清奈國際機場', 'type': 'international'},
    '加爾各答': {'lat': 22.6520, 'lon': 88.4463, 'name': '內塔吉蘇巴斯錢德拉鮑斯國際機場', 'type': 'international'},
    '海德拉巴': {'lat': 17.2313, 'lon': 78.4298, 'name': '拉吉夫甘地國際機場', 'type': 'international'},
    
    # 亞洲其他國家
    '沙烏地': {'lat': 24.9574, 'lon': 46.6988, 'name': '利雅得哈立德國王國際機場', 'type': 'international'},
    '科威特': {'lat': 29.2267, 'lon': 47.9689, 'name': '科威特國際機場', 'type': 'international'},
    '約旦': {'lat': 31.7225, 'lon': 35.9932, 'name': '安曼阿利亞王后國際機場', 'type': 'international'},
    '黎巴嫩': {'lat': 33.8208, 'lon': 35.4883, 'name': '貝魯特拉菲克哈里里國際機場', 'type': 'international'},
    '巴基斯坦': {'lat': 33.6167, 'lon': 73.0990, 'name': '伊斯蘭堡國際機場', 'type': 'international'},
    '孟加拉': {'lat': 23.8430, 'lon': 90.3979, 'name': '達卡沙賈拉拉國際機場', 'type': 'international'},
    '斯里蘭卡': {'lat': 7.1809, 'lon': 79.8842, 'name': '班達拉奈克國際機場', 'type': 'international'},
    '緬甸': {'lat': 16.9074, 'lon': 96.1332, 'name': '仰光國際機場', 'type': 'international'},
    '寮國': {'lat': 17.9883, 'lon': 102.5633, 'name': '永珍瓦岱國際機場', 'type': 'international'},
    '柬埔寨': {'lat': 11.5462, 'lon': 104.8442, 'name': '金邊國際機場', 'type': 'international'},
    '汶萊': {'lat': 4.9462, 'lon': 114.9283, 'name': '汶萊國際機場', 'type': 'international'},
    '不丹': {'lat': 27.5073, 'lon': 89.5249, 'name': '帕羅國際機場', 'type': 'international'},
    '蒙古': {'lat': 47.8439, 'lon': 106.7671, 'name': '成吉思汗國際機場', 'type': 'international'},
    '尼泊爾': {'lat': 27.6964, 'lon': 85.3590, 'name': '特里布万國際機場', 'type': 'international'},
    '阿富汗': {'lat': 34.5658, 'lon': 69.2125, 'name': '哈米德卡爾扎伊國際機場', 'type': 'international'},
    '伊朗': {'lat': 35.4161, 'lon': 51.1522, 'name': '伊瑪目霍梅尼國際機場', 'type': 'international'},
    '伊拉克': {'lat': 33.2625, 'lon': 44.2346, 'name': '巴格達國際機場', 'type': 'international'},
    '卡達': {'lat': 25.2731, 'lon': 51.6088, 'name': '哈馬德國際機場', 'type': 'international'},
    '巴林': {'lat': 26.2708, 'lon': 50.6336, 'name': '巴林國際機場', 'type': 'international'},
    '阿曼': {'lat': 23.5933, 'lon': 58.2844, 'name': '馬斯喀特國際機場', 'type': 'international'},
    '葉門': {'lat': 15.4766, 'lon': 44.2197, 'name': '薩那國際機場', 'type': 'international'},
    '敘利亞': {'lat': 33.4115, 'lon': 36.5157, 'name': '大馬士革國際機場', 'type': 'international'},
    '喬治亞': {'lat': 41.6692, 'lon': 44.9547, 'name': '提比里斯國際機場', 'type': 'international'},
    '亞美尼亞': {'lat': 40.1472, 'lon': 44.3959, 'name': '茲瓦爾特諾茨國際機場', 'type': 'international'},
    '亞塞拜然': {'lat': 40.4675, 'lon': 50.0467, 'name': '蓋達爾阿利耶夫國際機場', 'type': 'international'},
    '哈薩克': {'lat': 43.3547, 'lon': 77.0405, 'name': '阿拉木圖國際機場', 'type': 'international'},
    '吉爾吉斯': {'lat': 43.0613, 'lon': 74.4776, 'name': '瑪納斯國際機場', 'type': 'international'},
    '塔吉克': {'lat': 38.5439, 'lon': 68.8250, 'name': '杜尚別國際機場', 'type': 'international'},
    '土庫曼': {'lat': 37.9868, 'lon': 58.3610, 'name': '阿什哈巴德國際機場', 'type': 'international'},
    '烏茲別克': {'lat': 41.2578, 'lon': 69.2811, 'name': '塔什干國際機場', 'type': 'international'},
    
    # 日本主要城市
    '東京': {'lat': 35.5494, 'lon': 139.7798, 'name': '成田國際機場', 'type': 'international'},
    '大阪': {'lat': 34.4271, 'lon': 135.2442, 'name': '關西國際機場', 'type': 'international'},
    '札幌': {'lat': 42.7752, 'lon': 141.6922, 'name': '新千歲機場', 'type': 'international'},
    '福岡': {'lat': 33.5859, 'lon': 130.4508, 'name': '福岡機場', 'type': 'international'},
    '沖繩': {'lat': 26.1960, 'lon': 127.6460, 'name': '那霸機場', 'type': 'international'},
    '富山': {'lat': 36.6485, 'lon': 137.1878, 'name': '富山機場', 'type': 'international'},
    '藤澤': {'lat': 35.3390, 'lon': 139.4906, 'name': '羽田機場', 'type': 'international'},
    
    # 韓國主要城市
    '首爾': {'lat': 37.4602, 'lon': 126.4407, 'name': '仁川國際機場', 'type': 'international'},
    '釜山': {'lat': 35.1796, 'lon': 128.9382, 'name': '金海國際機場', 'type': 'international'},
    
    # 歐洲主要城市
    '巴黎': {'lat': 49.0097, 'lon': 2.5479, 'name': '戴高樂機場', 'type': 'international'},
    '倫敦': {'lat': 51.4700, 'lon': -0.4543, 'name': '希斯洛機場', 'type': 'international'},
    '柏林': {'lat': 52.3667, 'lon': 13.5033, 'name': '柏林勃蘭登堡機場', 'type': 'international'},
    '羅馬': {'lat': 41.8003, 'lon': 12.2389, 'name': '羅馬菲烏米奇諾機場', 'type': 'international'},
    '馬德里': {'lat': 40.4719, 'lon': -3.5626, 'name': '馬德里巴拉哈斯機場', 'type': 'international'},
    '阿姆斯特丹': {'lat': 52.3086, 'lon': 4.7639, 'name': '阿姆斯特丹史基浦機場', 'type': 'international'},
    '布魯塞爾': {'lat': 50.9014, 'lon': 4.4844, 'name': '布魯塞爾機場', 'type': 'international'},
    '波蘭': {'lat': 52.1657, 'lon': 20.9671, 'name': '華沙蕭邦機場', 'type': 'international'},
    '圓爾': {'lat': 49.0097, 'lon': 2.5479, 'name': '戴高樂機場', 'type': 'international'},
    '哥本哈根': {'lat': 55.6179, 'lon': 12.6560, 'name': '哥本哈根機場', 'type': 'international'},
    '里昂': {'lat': 45.7256, 'lon': 5.0811, 'name': '里昂聖埃克索佩里機場', 'type': 'international'},
    '卡地夫': {'lat': 51.3967, 'lon': -3.3433, 'name': '卡地夫機場', 'type': 'international'},
    '台地丹': {'lat': 53.4264, 'lon': -6.2499, 'name': '都柏林機場', 'type': 'international'},
    '都柏林': {'lat': 53.4264, 'lon': -6.2499, 'name': '都柏林機場', 'type': 'international'},
    '鹿特丹': {'lat': 51.9563, 'lon': 4.4378, 'name': '鹿特丹海牙機場', 'type': 'international'},
    
    # 歐洲其他國家
    '愛爾蘭': {'lat': 53.4264, 'lon': -6.2499, 'name': '都柏林機場', 'type': 'international'},
    '冰島': {'lat': 63.9850, 'lon': -22.6056, 'name': '凱夫拉維克國際機場', 'type': 'international'},
    '盧森堡': {'lat': 49.6233, 'lon': 6.2044, 'name': '盧森堡芬德爾機場', 'type': 'international'},
    '馬爾他': {'lat': 35.8575, 'lon': 14.4775, 'name': '馬爾他國際機場', 'type': 'international'},
    '克羅埃西亞': {'lat': 45.7429, 'lon': 16.0688, 'name': '薩格勒布機場', 'type': 'international'},
    '斯洛伐克': {'lat': 48.1702, 'lon': 17.2127, 'name': '布拉提斯拉瓦機場', 'type': 'international'},
    '保加利亞': {'lat': 42.6952, 'lon': 23.4064, 'name': '索菲亞機場', 'type': 'international'},
    '愛沙尼亞': {'lat': 59.4133, 'lon': 24.8328, 'name': '塔林機場', 'type': 'international'},
    '拉脫維亞': {'lat': 56.9236, 'lon': 23.9711, 'name': '里加國際機場', 'type': 'international'},
    '立陶宛': {'lat': 54.6339, 'lon': 25.2878, 'name': '維爾紐斯機場', 'type': 'international'},
    '烏克蘭': {'lat': 50.3450, 'lon': 30.8947, 'name': '鮑里斯波爾國際機場', 'type': 'international'},
    '白俄羅斯': {'lat': 53.8825, 'lon': 28.0308, 'name': '明斯克國際機場', 'type': 'international'},
    '摩爾多瓦': {'lat': 46.9276, 'lon': 28.9308, 'name': '基希涅夫國際機場', 'type': 'international'},
    '阿爾巴尼亞': {'lat': 41.4147, 'lon': 19.7206, 'name': '地拉那特蕾莎修女國際機場', 'type': 'international'},
    '北馬其頓': {'lat': 41.9616, 'lon': 21.6214, 'name': '史高比耶機場', 'type': 'international'},
    '波士尼亞': {'lat': 43.8246, 'lon': 18.3315, 'name': '塞拉耶佛國際機場', 'type': 'international'},
    '蒙特內哥羅': {'lat': 42.3594, 'lon': 19.2519, 'name': '波德戈里察機場', 'type': 'international'},
    '賽普勒斯': {'lat': 34.8751, 'lon': 33.6249, 'name': '拉納卡國際機場', 'type': 'international'},
    
    # 加拿大主要城市
    '多倫多': {'lat': 43.6777, 'lon': -79.6248, 'name': '多倫多機場', 'type': 'international'},
    '溫哥華': {'lat': 49.1947, 'lon': -123.1838, 'name': '溫哥華國際機場', 'type': 'international'},
    
    # 澳洲主要城市
    '昆士蘭': {'lat': -27.3818, 'lon': 153.1175, 'name': '布里斯本機場', 'type': 'international'},
    '凱恩斯': {'lat': -16.8742, 'lon': 145.7550, 'name': '凱恩斯機場', 'type': 'international'},
    '墨爾本': {'lat': -37.6690, 'lon': 144.8410, 'name': '墨爾本機場', 'type': 'international'},
    '巴塞隆納': {'lat': 41.2974, 'lon': 2.0833, 'name': '巴塞隆納機場', 'type': 'international'},
    
    # 東南亞主要城市
    '曼谷': {'lat': 13.6900, 'lon': 100.7501, 'name': '素萬那普機場', 'type': 'international'},
    '清邁': {'lat': 18.7696, 'lon': 98.9628, 'name': '清邁國際機場', 'type': 'international'},
    '吉隆坡': {'lat': 2.7456, 'lon': 101.7072, 'name': '吉隆坡國際機場', 'type': 'international'},
    '馬尼拉': {'lat': 14.5086, 'lon': 121.0198, 'name': '尼諾伊艾奎諾機場', 'type': 'international'},
    
    # 美洲其他國家
    '哥倫比亞': {'lat': 4.7016, 'lon': -74.1469, 'name': '艾爾多拉多國際機場', 'type': 'international'},
    '委內瑞拉': {'lat': 10.6012, 'lon': -66.9912, 'name': '西蒙玻利瓦爾國際機場', 'type': 'international'},
    '厄瓜多': {'lat': -0.1292, 'lon': -78.3575, 'name': '蘇克雷元帥國際機場', 'type': 'international'},
    '秘魯': {'lat': -12.0219, 'lon': -77.1143, 'name': '豪爾赫查維茲國際機場', 'type': 'international'},
    '玻利維亞': {'lat': -16.5113, 'lon': -68.1923, 'name': '艾爾阿爾托國際機場', 'type': 'international'},
    '巴拉圭': {'lat': -25.2398, 'lon': -57.5192, 'name': '亞松森國際機場', 'type': 'international'},
    '烏拉圭': {'lat': -34.8384, 'lon': -56.0308, 'name': '卡拉斯科國際機場', 'type': 'international'},
    '蓋亞那': {'lat': 6.4985, 'lon': -58.2543, 'name': '喬治敦國際機場', 'type': 'international'},
    '蘇利南': {'lat': 5.4531, 'lon': -55.1878, 'name': '約翰彭格爾國際機場', 'type': 'international'},
    '法屬圭亞那': {'lat': 4.8193, 'lon': -52.3603, 'name': '卡宴國際機場', 'type': 'international'},
    '哥斯大黎加': {'lat': 9.9938, 'lon': -84.2088, 'name': '胡安聖瑪麗亞國際機場', 'type': 'international'},
    '巴拿馬': {'lat': 9.0714, 'lon': -79.3835, 'name': '托庫門國際機場', 'type': 'international'},
    '尼加拉瓜': {'lat': 12.1415, 'lon': -86.1682, 'name': '奧古斯托桑地諾國際機場', 'type': 'international'},
    '宏都拉斯': {'lat': 14.0609, 'lon': -87.2172, 'name': '通孔廷國際機場', 'type': 'international'},
    '薩爾瓦多': {'lat': 13.4405, 'lon': -89.0557, 'name': '奧斯卡羅梅羅國際機場', 'type': 'international'},
    '瓜地馬拉': {'lat': 14.5833, 'lon': -90.5275, 'name': '拉奧羅拉國際機場', 'type': 'international'},
    '貝里斯': {'lat': 17.5395, 'lon': -88.3082, 'name': '菲利普戈德森國際機場', 'type': 'international'},
    '古巴': {'lat': 22.9892, 'lon': -82.4091, 'name': '何塞馬蒂國際機場', 'type': 'international'},
    '牙買加': {'lat': 18.5037, 'lon': -77.9137, 'name': '諾曼曼利國際機場', 'type': 'international'},
    '海地': {'lat': 18.5800, 'lon': -72.2925, 'name': '杜桑盧維杜爾國際機場', 'type': 'international'},
    '多明尼加': {'lat': 18.4297, 'lon': -69.6689, 'name': '美洲國際機場', 'type': 'international'},
    '波多黎各': {'lat': 18.4394, 'lon': -66.0018, 'name': '路易斯穆尼奧斯馬林國際機場', 'type': 'international'},
    '巴哈馬': {'lat': 25.0389, 'lon': -77.4663, 'name': '林登平德林國際機場', 'type': 'international'},
    '巴貝多': {'lat': 13.0747, 'lon': -59.4925, 'name': '格蘭特利亞當斯國際機場', 'type': 'international'},
    '千里達': {'lat': 10.5953, 'lon': -61.3372, 'name': '皮亞爾科國際機場', 'type': 'international'},
    
    # 非洲其他國家
    '奈及利亞': {'lat': 6.5774, 'lon': 3.3211, 'name': '穆爾塔拉穆罕默德國際機場', 'type': 'international'},
    '摩洛哥': {'lat': 33.3675, 'lon': -7.5898, 'name': '穆罕默德五世國際機場', 'type': 'international'},
    '突尼西亞': {'lat': 36.8510, 'lon': 10.2272, 'name': '突尼斯迦太基國際機場', 'type': 'international'},
    '阿爾及利亞': {'lat': 36.6910, 'lon': 3.2154, 'name': '阿爾及爾國際機場', 'type': 'international'},
    '利比亞': {'lat': 32.6635, 'lon': 13.1590, 'name': '的黎波里國際機場', 'type': 'international'},
    '蘇丹': {'lat': 15.5895, 'lon': 32.5532, 'name': '喀土穆國際機場', 'type': 'international'},
    '查德': {'lat': 12.1337, 'lon': 15.0340, 'name': '恩賈梅納國際機場', 'type': 'international'},
    '尼日': {'lat': 13.4815, 'lon': 2.1837, 'name': '迪奧里哈馬尼國際機場', 'type': 'international'},
    '馬利': {'lat': 12.5335, 'lon': -7.9499, 'name': '巴馬科國際機場', 'type': 'international'},
    '布吉納法索': {'lat': 12.3533, 'lon': -1.5122, 'name': '瓦加杜古機場', 'type': 'international'},
    '茅利塔尼亞': {'lat': 18.0983, 'lon': -15.9485, 'name': '努瓦克肖特國際機場', 'type': 'international'},
    '塞內加爾': {'lat': 14.7397, 'lon': -17.4903, 'name': '布萊茲迪亞涅國際機場', 'type': 'international'},
    '幾內亞': {'lat': 9.5769, 'lon': -13.6120, 'name': '科納克里國際機場', 'type': 'international'},
    '獅子山': {'lat': 8.6163, 'lon': -13.1956, 'name': '隆吉國際機場', 'type': 'international'},
    '賴比瑞亞': {'lat': 6.2338, 'lon': -10.3574, 'name': '羅伯茨國際機場', 'type': 'international'},
    '象牙海岸': {'lat': 5.2614, 'lon': -3.9263, 'name': '費利克斯烏弗埃博瓦尼機場', 'type': 'international'},
    '多哥': {'lat': 6.1656, 'lon': 1.2545, 'name': '洛美機場', 'type': 'international'},
    '貝南': {'lat': 6.3572, 'lon': 2.3844, 'name': '科托努機場', 'type': 'international'},
    '喀麥隆': {'lat': 3.7236, 'lon': 11.5534, 'name': '杜阿拉國際機場', 'type': 'international'},
    '中非': {'lat': 4.3985, 'lon': 18.5189, 'name': '班吉機場', 'type': 'international'},
    '赤道幾內亞': {'lat': 3.7553, 'lon': 8.7087, 'name': '馬拉博國際機場', 'type': 'international'},
    '加彭': {'lat': 0.4586, 'lon': 9.4123, 'name': '利伯維爾國際機場', 'type': 'international'},
    '剛果': {'lat': -4.2517, 'lon': 15.2530, 'name': '瑪雅瑪雅機場', 'type': 'international'},
    '剛果民主': {'lat': -4.3858, 'lon': 15.4445, 'name': '恩吉利國際機場', 'type': 'international'},
    '安哥拉': {'lat': -8.8583, 'lon': 13.2312, 'name': '羅安達國際機場', 'type': 'international'},
    '尚比亞': {'lat': -15.3308, 'lon': 28.4526, 'name': '肯尼思卡翁達國際機場', 'type': 'international'},
    '辛巴威': {'lat': -17.9318, 'lon': 31.0928, 'name': '哈拉雷國際機場', 'type': 'international'},
    '莫三比克': {'lat': -25.9208, 'lon': 32.5726, 'name': '馬普托國際機場', 'type': 'international'},
    '馬拉威': {'lat': -13.7894, 'lon': 33.7810, 'name': '利隆圭國際機場', 'type': 'international'},
    '波札那': {'lat': -24.5552, 'lon': 25.9184, 'name': '塞雷茨卡馬國際機場', 'type': 'international'},
    '納米比亞': {'lat': -22.4799, 'lon': 17.4709, 'name': '霍齊亞庫塔科國際機場', 'type': 'international'},
    '史瓦帝尼': {'lat': -26.3589, 'lon': 31.7165, 'name': '恩史瓦帝三世國際機場', 'type': 'international'},
    '賴索托': {'lat': -29.4619, 'lon': 27.5525, 'name': '莫舒舒一世國際機場', 'type': 'international'},
    '馬達加斯加': {'lat': -18.7969, 'lon': 47.4788, 'name': '伊瓦圖國際機場', 'type': 'international'},
    '模里西斯': {'lat': -20.4302, 'lon': 57.6836, 'name': '西沃薩古爾拉姆古蘭爵士國際機場', 'type': 'international'},
    '塞席爾': {'lat': -4.6742, 'lon': 55.5218, 'name': '塞席爾國際機場', 'type': 'international'},
    '葛摩': {'lat': -11.7108, 'lon': 43.2439, 'name': '賽義德易卜拉欣王子國際機場', 'type': 'international'},
    '吉布地': {'lat': 11.5474, 'lon': 43.1594, 'name': '吉布地國際機場', 'type': 'international'},
    '索馬利亞': {'lat': 2.0143, 'lon': 45.3047, 'name': '亞丁阿德國際機場', 'type': 'international'},
    '厄利垂亞': {'lat': 15.2919, 'lon': 38.9107, 'name': '阿斯馬拉國際機場', 'type': 'international'},
    '烏干達': {'lat': 0.0422, 'lon': 32.4435, 'name': '恩德培國際機場', 'type': 'international'},
    '盧安達': {'lat': -1.9686, 'lon': 30.1395, 'name': '基加利國際機場', 'type': 'international'},
    '蒲隆地': {'lat': -3.3239, 'lon': 29.3185, 'name': '梅爾希奧爾恩達達耶國際機場', 'type': 'international'},
    '南蘇丹': {'lat': 4.8720, 'lon': 31.6011, 'name': '朱巴國際機場', 'type': 'international'},
    
    # 大洋洲其他國家和地區
    '斐濟': {'lat': -17.7555, 'lon': 177.4436, 'name': '楠迪國際機場', 'type': 'international'},
    '巴布亞紐幾內亞': {'lat': -9.4438, 'lon': 147.2200, 'name': '傑克遜斯國際機場', 'type': 'international'},
    '索羅門群島': {'lat': -9.4280, 'lon': 160.0548, 'name': '霍尼亞拉國際機場', 'type': 'international'},
    '萬那杜': {'lat': -17.6993, 'lon': 168.3270, 'name': '鮑爾菲爾德國際機場', 'type': 'international'},
    '新喀里多尼亞': {'lat': -22.0146, 'lon': 166.2128, 'name': '努美阿國際機場', 'type': 'international'},
    '大溪地': {'lat': -17.5567, 'lon': -149.6068, 'name': '法阿國際機場', 'type': 'international'},
    '薩摩亞': {'lat': -13.8297, 'lon': -171.9967, 'name': '法萊奧洛國際機場', 'type': 'international'},
    '東加': {'lat': -21.2411, 'lon': -175.1498, 'name': '富阿阿莫圖國際機場', 'type': 'international'},
    '吉里巴斯': {'lat': 1.9863, 'lon': -157.3500, 'name': '邦里基國際機場', 'type': 'international'},
    '馬紹爾群島': {'lat': 7.0647, 'lon': 171.2720, 'name': '阿馬塔卡布阿國際機場', 'type': 'international'},
    '密克羅尼西亞': {'lat': 6.9851, 'lon': 158.2090, 'name': '波納佩國際機場', 'type': 'international'},
    '諾魯': {'lat': -0.5477, 'lon': 166.9191, 'name': '諾魯國際機場', 'type': 'international'},
    '吐瓦魯': {'lat': -8.5250, 'lon': 179.1956, 'name': '富納富提國際機場', 'type': 'international'},
    '關島': {'lat': 13.4840, 'lon': 144.7970, 'name': '安東尼奧汪帕特國際機場', 'type': 'international'},
    '塞班': {'lat': 15.1190, 'lon': 145.7296, 'name': '塞班國際機場', 'type': 'international'},
    
    # 其他特殊地點
    '儀賢會': {'lat': 25.0356, 'lon': 121.4320, 'name': '輔仁大學', 'type': 'domestic'},
    '大板根酒店': {'lat': 24.9000, 'lon': 121.4000, 'name': '大板根森林溫泉渡假村', 'type': 'domestic'},
    '璞園': {'lat': 24.9970, 'lon': 121.2620, 'name': '璞園訓練中心', 'type': 'domestic'},
    '陽明山溫泉渡假村': {'lat': 25.1505, 'lon': 121.5363, 'name': '陽明山溫泉區', 'type': 'domestic'},
    '沙鹿翔元會館': {'lat': 24.2377, 'lon': 120.5579, 'name': '台中沙鹿區', 'type': 'domestic'},
    '香草星空民宿': {'lat': 24.8593, 'lon': 121.8230, 'name': '宜蘭頭城', 'type': 'domestic'},
}

# 別名對照表（支援不同的輸入方式）
ALIASES = {
    # 台北相關
    'taipei': '台北',
    'Taipei': '台北',
    'TAIPEI': '台北',
    '台北市': '台北',
    '臺北市': '台北',
    'tp': '台北',
    'TPE': '台北',
    
    # 新北相關
    'new taipei': '新北',
    'New Taipei': '新北',
    'newtaipei': '新北',
    '新北市': '新北',
    'NTPC': '新北',
    'ntpc': '新北',
    
    # 桃園相關
    'taoyuan': '桃園',
    'Taoyuan': '桃園',
    '桃園市': '桃園',
    '桃園縣': '桃園',
    'ty': '桃園',
    
    # 台中相關
    'taichung': '台中',
    'Taichung': '台中',
    '台中市': '台中',
    '臺中市': '台中',
    'tc': '台中',
    
    # 台南相關
    'tainan': '台南',
    'Tainan': '台南',
    '台南市': '台南',
    '臺南市': '台南',
    'tn': '台南',
    
    # 高雄相關
    'kaohsiung': '高雄',
    'Kaohsiung': '高雄',
    '高雄市': '高雄',
    'ks': '高雄',
    'kh': '高雄',
    
    # 新竹相關
    'hsinchu': '新竹',
    'Hsinchu': '新竹',
    '新竹市': '新竹',
    '新竹縣': '新竹',
    'hc': '新竹',
    
    # 彰化相關
    'changhua': '彰化',
    'Changhua': '彰化',
    '彰化市': '彰化',
    '彰化縣': '彰化',
    'ch': '彰化',
    
    # 雲林相關
    'yunlin': '雲林',
    'Yunlin': '雲林',
    '雲林縣': '雲林',
    'yl': '雲林',
    
    # 苗栗相關
    'miaoli': '苗栗',
    'Miaoli': '苗栗',
    '苗栗市': '苗栗',
    '苗栗縣': '苗栗',
    'ml': '苗栗',
    
    # 嘉義相關
    'chiayi': '嘉義',
    'Chiayi': '嘉義',
    '嘉義市': '嘉義',
    '嘉義縣': '嘉義',
    'cy': '嘉義',
    
    # 屏東相關
    'pingtung': '屏東',
    'Pingtung': '屏東',
    '屏東市': '屏東',
    '屏東縣': '屏東',
    'pt': '屏東',
    
    # 花蓮相關
    'hualien': '花蓮',
    'Hualien': '花蓮',
    '花蓮市': '花蓮',
    '花蓮縣': '花蓮',
    'hl': '花蓮',
    
    # 台東相關
    'taitung': '台東',
    'Taitung': '台東',
    '台東市': '台東',
    '台東縣': '台東',
    '臺東市': '台東',
    '臺東縣': '台東',
    'tt': '台東',
    
    # 宜蘭相關
    'yilan': '宜蘭',
    'Yilan': '宜蘭',
    '宜蘭市': '宜蘭',
    '宜蘭縣': '宜蘭',
    'yl': '宜蘭',
    '宜蘭': '宜蘭',
    
    # 基隆相關
    'keelung': '基隆',
    'Keelung': '基隆',
    '基隆市': '基隆',
    'kl': '基隆',
    
    # 南投相關
    'nantou': '南投',
    'Nantou': '南投',
    '南投市': '南投',
    '南投縣': '南投',
    'nt': '南投',
    
    # 澎湖相關
    'penghu': '澎湖',
    'Penghu': '澎湖',
    '澎湖縣': '澎湖',
    'ph': '澎湖',
    
    # 金門相關
    'kinmen': '金門',
    'Kinmen': '金門',
    '金門縣': '金門',
    'km': '金門',
    
    # 馬祖相關
    'matsu': '馬祖',
    'Matsu': '馬祖',
    '連江縣': '馬祖',
    'mt': '馬祖',
    
    # 國際地點
    'japan': '日本',
    'Japan': '日本',
    'JP': '日本',
    'jp': '日本',
    '日': '日本',
    
    'korea': '韓國',
    'Korea': '韓國',
    'KR': '韓國',
    'kr': '韓國',
    'south korea': '韓國',
    'South Korea': '韓國',
    '南韓': '韓國',
    
    'singapore': '新加坡',
    'Singapore': '新加坡',
    'SG': '新加坡',
    'sg': '新加坡',
    '星島': '新加坡',
    '星洲': '新加坡',
    
    'thailand': '泰國',
    'Thailand': '泰國',
    'TH': '泰國',
    'th': '泰國',
    '泰': '泰國',
    
    'usa': '美國',
    'USA': '美國',
    'US': '美國',
    'us': '美國',
    'america': '美國',
    'America': '美國',
    'United States': '美國',
    '美': '美國',
    
    'germany': '德國',
    'Germany': '德國',
    'DE': '德國',
    'de': '德國',
    '德': '德國',
    
    'uk': '英國',
    'UK': '英國',
    'GB': '英國',
    'gb': '英國',
    'britain': '英國',
    'Britain': '英國',
    'england': '英國',
    'England': '英國',
    'United Kingdom': '英國',
    '英': '英國',
    
    'france': '法國',
    'France': '法國',
    'FR': '法國',
    'fr': '法國',
    '法': '法國',
    
    'australia': '澳洲',
    'Australia': '澳洲',
    'AU': '澳洲',
    'au': '澳洲',
    '澳': '澳洲',
    '澳大利亞': '澳洲',
    
    'canada': '加拿大',
    'Canada': '加拿大',
    'CA': '加拿大',
    'ca': '加拿大',
    '加': '加拿大',
    
    # 中國地點
    'china': '中國',
    'China': '中國',
    'CN': '中國',
    'cn': '中國',
    '大陸': '中國',
    '中國大陸': '中國',
    
    'shanghai': '上海',
    'Shanghai': '上海',
    '上海市': '上海',
    'sh': '上海',
    
    'beijing': '中國',
    'Beijing': '中國',
    '北京': '中國',
    '北京市': '中國',
    'bj': '中國',
    
    'guangzhou': '廣州',
    'Guangzhou': '廣州',
    '廣州市': '廣州',
    'gz': '廣州',
    '穗': '廣州',
    
    'shenzhen': '深圳',
    'Shenzhen': '深圳',
    '深圳市': '深圳',
    'sz': '深圳',
    '深': '深圳',
    
    'hangzhou': '杭州',
    'Hangzhou': '杭州',
    '杭州市': '杭州',
    'hz': '杭州',
    
    'nanjing': '南京',
    'Nanjing': '南京',
    '南京市': '南京',
    'nj': '南京',
    
    'chengdu': '成都',
    'Chengdu': '成都',
    '成都市': '成都',
    'cd': '成都',
    
    'xian': '西安',
    'Xian': '西安',
    "xi'an": '西安',
    "Xi'an": '西安',
    '西安市': '西安',
    'xa': '西安',
    
    'wuhan': '武漢',
    'Wuhan': '武漢',
    '武漢市': '武漢',
    'wh': '武漢',
    
    'xiamen': '廈門',
    'Xiamen': '廈門',
    '廈門市': '廈門',
    'xm': '廈門',
    '廈': '廈門',
    
    'qingdao': '青島',
    'Qingdao': '青島',
    '青島市': '青島',
    'qd': '青島',
    
    'dalian': '大連',
    'Dalian': '大連',
    '大連市': '大連',
    'dl': '大連',
    
    'chongqing': '重慶',
    'Chongqing': '重慶',
    '重慶市': '重慶',
    'cq': '重慶',
    
    'tianjin': '天津',
    'Tianjin': '天津',
    '天津市': '天津',
    'tj': '天津',
    
    # 大學簡稱
    '清大': '清華大學',
    '成大': '成功大學',
    '高醫': '高醫大',
    '輔大': '輔仁大學',
    '嘉大': '嘉義大學',
    '台南成大': '成功大學',
    '成功大學': '成功',
    '高雄樹德科大': '樹德科技大學',
    '樹德科技大學': '樹德科大',
    '基隆海洋大學': '海洋',
    '臺科大': '台科大',
    '佛光': '佛光大學',
    
    # 其他地點別名
    '大板根森林': '大板根',
    '大板根渡假村': '大板根',
    '劍潭': '劍潭青年活動中心',
    '龍門': '龍門露營區',
    '臺中榮總': '台中榮民總醫院',
    '台中榮總': '台中榮民總醫院',
    '羅東聖母': '羅東聖母醫院',
}

# 台灣地區對應關係（用於判斷國內地點）
TAIWAN_REGION_MAPPING = {
    # 台北地區對應
    '南港': '台北',
    '松山': '台北',
    '信義': '台北',
    '大安': '台北',
    '中山': '台北',
    '內湖': '台北',
    '士林': '台北',
    '北投': '台北',
    '萬華': '台北',
    '中正': '台北',
    '大同': '台北',
    '文山': '台北',
    '大直': '台北',
    '關渡': '台北',
    
    # 新北地區對應
    '板橋': '新北',
    '三重': '新北',
    '中和': '新北',
    '永和': '新北',
    '新店': '新北',
    '汐止': '新北',
    '淡水': '新北',
    '土城': '新北',
    '蘆洲': '新北',
    '樹林': '新北',
    '三峽': '新北',
    '鶯歌': '新北',
    '深坑': '新北',
    '石碇': '新北',
    '坪林': '新北',
    '貢寮': '新北',
    '雙溪': '新北',
    '瑞芳': '新北',
    '平溪': '新北',
    '烏來': '新北',
    '林口': '新北',
    '五股': '新北',
    '泰山': '新北',
    '八里': '新北',
    '金山': '新北',
    '萬里': '新北',
}

# 國際城市對應關係（用於找對應機場）
INTERNATIONAL_CITY_MAPPING = {
    # 中國城市對應機場
    '蘇州': '上海',  # 蘇州使用上海機場
    '無錫': '上海',  # 無錫使用上海機場
    '寧波': '杭州',  # 寧波使用杭州機場
    '佛山': '廣州',  # 佛山使用廣州機場
    '東莞': '深圳',  # 東莞使用深圳機場
    '珠海': '廣州',  # 珠海使用廣州機場
    
    # 國際城市對應到國家
    '東京': '日本',
    '大阪': '日本',
    '京都': '日本',
    '名古屋': '日本',
    '福岡': '日本',
    '札幌': '日本',
    '首爾': '韓國',
    '釜山': '韓國',
    '仁川': '韓國',
    '曼谷': '泰國',
    '清邁': '泰國',
    '普吉': '泰國',
    '吉隆坡': '馬來西亞',
    '檳城': '馬來西亞',
    '胡志明': '越南',
    '河內': '越南',
    '雅加達': '印尼',
    '峇里島': '印尼',
    '馬尼拉': '菲律賓',
    '宿霧': '菲律賓',
    '新德里': '印度',
    '孟買': '印度',
    '倫敦': '英國',
    '曼徹斯特': '英國',
    '巴黎': '法國',
    '馬賽': '法國',
    '法蘭克福': '德國',
    '慕尼黑': '德國',
    '柏林': '德國',
    '紐約': '美國',
    '洛杉磯': '美國',
    '舊金山': '美國',
    '芝加哥': '美國',
    '西雅圖': '美國',
    '波士頓': '美國',
    '華盛頓': '美國',
    '多倫多': '加拿大',
    '溫哥華': '加拿大',
    '雪梨': '澳洲',
    '墨爾本': '澳洲',
    '布里斯本': '澳洲',
    '奧克蘭': '紐西蘭',
    '威靈頓': '紐西蘭',
}

# 中國城市關鍵字（用於判斷是否為中國城市）
CHINA_KEYWORDS = ['北京', '上海', '廣州', '深圳', '杭州', '成都', '西安', '武漢', 
                  '廈門', '南京', '青島', '大連', '重慶', '天津', '鄭州', '長沙',
                  '昆明', '福州', '蘇州', '無錫', '寧波', '佛山', '東莞', '珠海',
                  '合肥', '濟南', '哈爾濱', '瀋陽', '長春', '石家莊', '太原',
                  '南昌', '貴陽', '海口', '蘭州', '銀川', '西寧', '烏魭木齊',
                  '廚州', '南寧', '無錫', '溫州', '嘉興', '紹興', '南通', '常州',
                  '徐州', '揚州', '泉州', '漳州', '莵田', '煙台', '濰州', '威海']

# ========== 碳排放 / 縮放：優先使用可測試套件 ==========
try:
    from travel_carbon.carbon import (  # type: ignore
        CARBON_EMISSION_FACTORS,
        determine_transport_mode,
        calculate_carbon_emission,
    )
    from travel_carbon.zoom import calculate_zoom_level  # type: ignore
except ImportError:
    # Fallback when package not installed (python path = repo root only)
    # 簡化係數 (kg CO2e/km)；正式盤查請對照環境部公告，見 docs/emission_factors_notes.md
    CARBON_EMISSION_FACTORS = {
        '自用車/計程車': 0.21,
        '國內航班': 0.133,
        '國際航班': 0.101,
        '火車/高鐵': 0.034,
        '巴士': 0.065,
    }

    def determine_transport_mode(travel_type, distance_km):
        if distance_km <= 0:
            return ('N/A', 0)
        if '國際' in travel_type:
            return ('國際航班', CARBON_EMISSION_FACTORS['國際航班'])
        elif '離島' in travel_type or '航班' in travel_type:
            return ('國內航班', CARBON_EMISSION_FACTORS['國內航班'])
        elif '開車' in travel_type:
            if distance_km > 300:
                return ('火車/高鐵', CARBON_EMISSION_FACTORS['火車/高鐵'])
            else:
                return ('自用車/計程車', CARBON_EMISSION_FACTORS['自用車/計程車'])
        else:
            if distance_km > 500:
                return ('國內航班', CARBON_EMISSION_FACTORS['國內航班'])
            elif distance_km > 300:
                return ('火車/高鐵', CARBON_EMISSION_FACTORS['火車/高鐵'])
            else:
                return ('自用車/計程車', CARBON_EMISSION_FACTORS['自用車/計程車'])

    def calculate_carbon_emission(travel_type, distance_km):
        transport_mode, factor = determine_transport_mode(travel_type, distance_km)
        emission = round(distance_km * factor, 3) if distance_km > 0 else 0
        return {
            '交通方式': transport_mode,
            '碳排係數': factor,
            '碳排放量(kg CO2e)': emission,
        }

    def calculate_zoom_level(distance_km):
        if distance_km < 10:
            return 13
        elif distance_km < 20:
            return 12
        elif distance_km < 40:
            return 11
        elif distance_km < 80:
            return 10
        elif distance_km < 150:
            return 9
        elif distance_km < 300:
            return 8
        elif distance_km < 600:
            return 7
        elif distance_km < 1500:
            return 6
        elif distance_km < 3000:
            return 5
        elif distance_km < 6000:
            return 4
        else:
            return 3
# ========== 碳排放 / 縮放結束 ==========

class TravelDistanceCalculator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("出差距離計算與地圖產生工具 (效率優化版)")
        self.root.geometry("1200x800")
        
        # 確保視窗在最前面
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(self.root.attributes, '-topmost', False)
        
        # 設定樣式
        style = ttk.Style()
        style.theme_use('clam')
        
        # 變數
        self.input_file = tk.StringVar()
        self.output_dir = tk.StringVar(value=os.path.expanduser("~/Downloads/差旅費"))
        self.start_row = tk.IntVar(value=1)
        self.end_row = tk.IntVar(value=10000)  # 預設處理到第10000列
        self.progress = tk.DoubleVar()
        self.status_text = tk.StringVar(value="準備就緒")
        
        # 資料存儲
        self.df_locations = None
        self.results = []
        self.unique_routes = {}  # 儲存唯一路線資訊
        
        # 快取機制
        self.route_cache = {}  # 儲存已計算的路線
        self.map_cache = {}    # 儲存已產生的地圖檔案
        self.load_cache()      # 載入既有快取
        
        self.setup_ui()
        
    def load_cache(self):
        """載入快取檔案"""
        cache_dir = os.path.expanduser("~/Downloads/差旅費/.cache")
        os.makedirs(cache_dir, exist_ok=True)
        
        # 載入路線快取
        route_cache_file = os.path.join(cache_dir, 'route_cache.pkl')
        if os.path.exists(route_cache_file):
            try:
                with open(route_cache_file, 'rb') as f:
                    self.route_cache = pickle.load(f)
                print(f"載入 {len(self.route_cache)} 條快取路線")
            except:
                self.route_cache = {}
                
        # 載入地圖快取
        map_cache_file = os.path.join(cache_dir, 'map_cache.json')
        if os.path.exists(map_cache_file):
            try:
                with open(map_cache_file, 'r', encoding='utf-8') as f:
                    self.map_cache = json.load(f)
                # 檢查快取的地圖檔案是否仍存在
                valid_cache = {}
                for key, path in self.map_cache.items():
                    if os.path.exists(path):
                        valid_cache[key] = path
                self.map_cache = valid_cache
                print(f"載入 {len(self.map_cache)} 個有效快取地圖")
            except:
                self.map_cache = {}
                
    def save_cache(self):
        """儲存快取檔案"""
        cache_dir = os.path.expanduser("~/Downloads/差旅費/.cache")
        
        # 儲存路線快取
        route_cache_file = os.path.join(cache_dir, 'route_cache.pkl')
        with open(route_cache_file, 'wb') as f:
            pickle.dump(self.route_cache, f)
            
        # 儲存地圖快取
        map_cache_file = os.path.join(cache_dir, 'map_cache.json')
        with open(map_cache_file, 'w', encoding='utf-8') as f:
            json.dump(self.map_cache, f, ensure_ascii=False, indent=2)
        
    def setup_ui(self):
        """設置UI介面"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 檔案選擇區
        file_frame = ttk.LabelFrame(main_frame, text="檔案設定", padding="10")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(file_frame, text="輸入檔案:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(file_frame, textvariable=self.input_file, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(file_frame, text="瀏覽", command=self.browse_input_file).grid(row=0, column=2)
        
        ttk.Label(file_frame, text="輸出目錄:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(file_frame, textvariable=self.output_dir, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(file_frame, text="瀏覽", command=self.browse_output_dir).grid(row=1, column=2)
        
        # 處理範圍設定
        range_frame = ttk.LabelFrame(main_frame, text="處理範圍", padding="10")
        range_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(range_frame, text="起始列:").grid(row=0, column=0, sticky=tk.W)
        ttk.Spinbox(range_frame, from_=1, to=10000, textvariable=self.start_row, width=10).grid(row=0, column=1, padx=5)
        
        ttk.Label(range_frame, text="結束列:").grid(row=0, column=2, sticky=tk.W, padx=(20, 0))
        ttk.Spinbox(range_frame, from_=1, to=10000, textvariable=self.end_row, width=10).grid(row=0, column=3, padx=5)
        
        # 快取資訊
        self.cache_info_label = ttk.Label(range_frame, text=f"快取: {len(self.route_cache)} 條路線, {len(self.map_cache)} 個地圖")
        self.cache_info_label.grid(row=0, column=4, padx=(20, 0))
        
        # 效率資訊
        self.efficiency_label = ttk.Label(range_frame, text="")
        self.efficiency_label.grid(row=1, column=0, columnspan=5, pady=5)
        
        # 控制按鈕
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="載入資料", command=self.load_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="計算距離", command=self.calculate_distances).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="產生地圖", command=self.generate_maps).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="匯出Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="清除快取", command=self.clear_cache).pack(side=tk.LEFT, padx=5)
        
        # 新增一鍵處理全部按鈕
        ttk.Separator(button_frame, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)
        self.process_all_btn = ttk.Button(button_frame, text="一鍵處理全部", 
                                         command=self.process_all_data, 
                                         style='Accent.TButton')
        self.process_all_btn.pack(side=tk.LEFT, padx=5)
        
        # 進度條
        progress_frame = ttk.Frame(main_frame)
        progress_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress, maximum=100)
        self.progress_bar.pack(fill=tk.X, expand=True)
        
        # 狀態列
        status_frame = ttk.Frame(main_frame)
        status_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        ttk.Label(status_frame, textvariable=self.status_text).pack(side=tk.LEFT)
        
        # 結果顯示區
        result_frame = ttk.LabelFrame(main_frame, text="處理結果", padding="10")
        result_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # 建立Treeview
        columns = ('序號', '傳票編號', '目的地', '類型', '距離(km)', '時間', '地圖', '快取')
        self.tree = ttk.Treeview(result_frame, columns=columns, show='headings', height=15)
        
        # 設定欄位
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100 if col != '快取' else 60)
        
        # 捲軸
        scrollbar = ttk.Scrollbar(result_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 配置grid權重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(5, weight=1)
        
    def update_cache_info(self):
        """更新快取資訊顯示"""
        self.cache_info_label.config(text=f"快取: {len(self.route_cache)} 條路線, {len(self.map_cache)} 個地圖")
        
    def process_all_data(self):
        """一鍵處理所有資料"""
        if self.df_locations is None:
            messagebox.showerror("錯誤", "請先載入資料")
            return
            
        # 確認處理所有資料
        total_rows = len(self.df_locations)
        if messagebox.askyesno("確認", f"確定要處理所有 {total_rows} 筆資料嗎？\n\n這可能需要一些時間。"):
            # 設定處理範圍為全部
            self.start_row.set(1)
            self.end_row.set(total_rows)
            
            # 依序執行：計算距離 -> 產生地圖 -> 匯出Excel
            self.status_text.set("開始一鍵處理...")
            
            # 在新執行緒中執行
            thread = threading.Thread(target=self._process_all_thread)
            thread.start()
            
    def _process_all_thread(self):
        """一鍵處理的執行緒函數"""
        try:
            # 步驟1: 計算距離
            self.root.after(0, lambda: self.status_text.set("步驟 1/3: 計算距離..."))
            self._calculate_distances_thread()
            
            # 等待計算完成
            while self.progress.get() < 100:
                time.sleep(0.1)
            
            # 步驟2: 產生地圖
            self.root.after(0, lambda: self.status_text.set("步驟 2/3: 產生地圖..."))
            self.progress.set(0)  # 重置進度條
            self._generate_maps_thread()
            
            # 等待地圖產生完成
            while self.progress.get() < 100:
                time.sleep(0.1)
            
            # 步驟3: 匯出Excel
            self.root.after(0, lambda: self.status_text.set("步驟 3/3: 匯出Excel..."))
            self.root.after(0, self.export_excel)
            
            # 完成
            self.root.after(0, lambda: messagebox.showinfo("完成", 
                f"一鍵處理完成！\n\n" +
                f"已處理 {len(self.results)} 筆資料\n" +
                f"產生 {len(self.unique_routes)} 個唯一路線地圖\n" +
                f"結果已匯出至 Excel 檔案"))
                
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("錯誤", f"一鍵處理失敗: {str(e)}"))
            self.root.after(0, lambda: self.status_text.set("一鍵處理失敗"))
        
    def clear_cache(self):
        """清除快取"""
        if messagebox.askyesno("確認", "確定要清除所有快取嗎？"):
            self.route_cache = {}
            self.map_cache = {}
            self.save_cache()
            self.update_cache_info()
            messagebox.showinfo("成功", "快取已清除")
            self.status_text.set("快取已清除")
        
    def browse_input_file(self):
        """選擇輸入檔案"""
        filename = filedialog.askopenfilename(
            title="選擇Excel檔案",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.input_file.set(filename)
            
    def browse_output_dir(self):
        """選擇輸出目錄"""
        directory = filedialog.askdirectory(title="選擇輸出目錄")
        if directory:
            self.output_dir.set(directory)
            
    def load_data(self):
        """載入Excel資料"""
        if not self.input_file.get():
            messagebox.showerror("錯誤", "請先選擇輸入檔案")
            return
            
        try:
            self.status_text.set("正在載入資料...")
            self.df_locations = pd.read_excel(self.input_file.get())
            
            # 顯示資料資訊
            total_rows = len(self.df_locations)
            
            # 分析資料內容
            destinations_col = '出差地點' if '出差地點' in self.df_locations.columns else None
            if destinations_col:
                # 計算有出差地點的資料筆數
                has_destination = self.df_locations[destinations_col].notna().sum()
                no_destination = total_rows - has_destination
                info_msg = f"成功載入 {total_rows} 筆資料\n\n" + \
                          f"- 有出差地點: {has_destination} 筆\n" + \
                          f"- 無出差地點: {no_destination} 筆"
            else:
                info_msg = f"成功載入 {total_rows} 筆資料\n\n注意：找不到'出差地點'欄位"
                
            messagebox.showinfo("載入成功", info_msg)
            self.status_text.set(f"已載入 {total_rows} 筆資料")
            
            # 更新範圍設定為全部資料
            self.end_row.set(total_rows)  # 預設處理所有資料
            
        except Exception as e:
            messagebox.showerror("錯誤", f"載入資料失敗: {str(e)}")
            self.status_text.set("載入失敗")
            
    def calculate_distances(self):
        """計算距離"""
        if self.df_locations is None:
            messagebox.showerror("錯誤", "請先載入資料")
            return
            
        # 在新執行緒中執行計算
        thread = threading.Thread(target=self._calculate_distances_thread)
        thread.start()
        
    def _calculate_distances_thread(self):
        """計算距離的執行緒函數"""
        try:
            start_idx = self.start_row.get() - 1
            end_idx = self.end_row.get()
            
            # 清空結果
            self.results = []
            self.unique_routes = {}
            self.tree.delete(*self.tree.get_children())
            
            # 處理選定範圍的資料
            subset = self.df_locations.iloc[start_idx:end_idx]
            total = len(subset)
            
            # 先分析所有唯一的目的地
            destinations_count = defaultdict(int)
            for _, row in subset.iterrows():
                dests = str(row.get('出差地點', '')).split('、') if pd.notna(row.get('出差地點', '')) else []
                if dests and dests[0]:
                    destinations_count[dests[0]] += 1
                    
            # 顯示效率資訊
            unique_count = len(destinations_count)
            self.root.after(0, lambda: self.efficiency_label.config(
                text=f"發現 {unique_count} 個不同目的地，共 {total} 筆資料"
            ))
            
            cache_hit = 0
            api_calls = 0
            skipped = 0  # 記錄跳過的筆數
            
            for idx, (_, row) in enumerate(subset.iterrows()):
                self.status_text.set(f"正在處理第 {idx+1}/{total} 筆...")
                self.progress.set((idx + 1) / total * 100)
                
                # 取得目的地
                destinations = str(row.get('出差地點', '')).split('、') if pd.notna(row.get('出差地點', '')) else []
                
                # 處理特殊情況
                if destinations and destinations[0] and destinations[0].strip() not in ['無地點資訊', '無', '']:
                    destination = destinations[0].strip()  # 取第一個目的地並去除空白
                    
                    # 智能處理目的地
                    destination = self._smart_destination_handler(destination)
                    
                    # 特殊處理離島（使用飛機）
                    if destination in ['澎湖', '金門', '馬祖'] or (destination in COORDINATES and COORDINATES[destination].get('type') == 'island'):
                        # 台灣離島：使用飛機
                        result = self._calculate_island_flight_distance(row, destination)
                        if result and result.get('cached'):
                            cache_hit += 1
                    elif self._is_domestic(destination):
                        # 台灣本島：統一從輔大出發開車
                        result = self._calculate_driving_distance(row, destination)
                        if result and result.get('cached'):
                            cache_hit += 1
                        else:
                            api_calls += 1
                    else:
                        # 國外：計算飛行距離（從桃園機場出發）
                        result = self._calculate_flight_distance(row, destination)
                        if result and result.get('cached'):
                            cache_hit += 1
                    
                    if result:
                        self.results.append(result)
                        # 更新UI
                        self.root.after(0, self._add_result_to_tree, result)
                        
                        # 收集唯一路線資訊
                        route_key = result.get('route_key')
                        if route_key and route_key not in self.unique_routes:
                            self.unique_routes[route_key] = {
                                '起點': result['起點'],
                                '終點': result['終點'],
                                '類型': result['類型'],
                                '距離(km)': result['距離(km)'],
                                '時間': result['時間'],
                                'route': result.get('route'),
                                'results': []
                            }
                        if route_key:
                            self.unique_routes[route_key]['results'].append(result)
                else:
                    # 如果沒有出差地點，仍然保留這筆記錄
                    # 檢查原始資料內容
                    original_location = str(row.get('出差地點', ''))
                    if original_location in ['無地點資訊', '無']:
                        display_text = original_location
                    else:
                        display_text = '無出差地點'
                        
                    empty_result = {
                        '序號': row['序號'],
                        '傳票編號': row['傳票編號'],
                        '目的地': display_text,
                        '類型': 'N/A',
                        '距離(km)': 0,
                        '時間': 'N/A',
                        '起點': {'name': '輔仁大學', 'lat': 25.0356, 'lon': 121.4320},
                        '終點': {'name': '無', 'lat': 0, 'lon': 0},
                        'route': None,
                        'route_key': 'no-destination',
                        'cached': False
                    }
                    self.results.append(empty_result)
                    self.root.after(0, self._add_result_to_tree, empty_result)
                
                # 減少API請求頻率
                if api_calls > 0 and (api_calls % 5 == 0):
                    time.sleep(1)
                    
            # 儲存快取
            self.save_cache()
            self.update_cache_info()
            
            # 統計處理結果
            processed = len([r for r in self.results if r.get('距離(km)', 0) > 0])
            no_destination = len([r for r in self.results if r.get('目的地') == '無出差地點'])
            unknown = len([r for r in self.results if '未知' in r.get('類型', '')])
            
            self.status_text.set(
                f"計算完成！共 {len(self.results)} 筆 "
                f"(成功處理 {processed} 筆，無目的地 {no_destination} 筆，未知地點 {unknown} 筆)，"
                f"{len(self.unique_routes)} 條唯一路線，"
                f"快取命中 {cache_hit} 筆，新計算 {api_calls} 筆"
            )
            self.progress.set(100)
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("錯誤", f"計算失敗: {str(e)}"))
            self.status_text.set("計算失敗")
            
    def _smart_destination_handler(self, destination):
        """智能處理目的地名稱"""
        original = destination
        
        # 先檢查別名對照表
        if destination.lower() in ALIASES:
            destination = ALIASES[destination.lower()]
        elif destination in ALIASES:
            destination = ALIASES[destination]
        
        # 移除常見的地名後綴
        suffixes_to_remove = ['市', '縣', '區', '鎮', '鄉', '街', '路', '段']
        for suffix in suffixes_to_remove:
            if destination.endswith(suffix):
                destination = destination[:-len(suffix)]
                
        # 處理特殊情況
        # 例如："台北市南港區" -> "南港"
        for region, parent in TAIWAN_REGION_MAPPING.items():
            if region in destination and parent in destination:
                return region
                
        # 處理中國城市的省份問題
        # 例如："浙江省杭州市" -> "杭州"
        china_provinces = ['浙江', '江蘇', '廣東', '福建', '山東', '河北', '河南', 
                          '湖北', '湖南', '四川', '陕西', '山西', '遼寧', '吉林', 
                          '黑龍江', '安徽', '江西', '雲南', '貴州', '甘肅', '青海']
        
        for province in china_provinces:
            if province in original:  # 使用原始字串
                # 找出省份後的城市名
                for city in CHINA_KEYWORDS:
                    if city in original:
                        return city
                        
        # 處理國際城市
        # 檢查是否在國際城市對應中
        if destination in INTERNATIONAL_CITY_MAPPING:
            return destination  # 保持原樣，讓後續處理
                        
        return destination
            
    def _is_domestic(self, destination):
        """判斷是否為國內地點（台灣本島用開車，離島用飛機）"""
        # 台灣本島縣市列表（統一從輔大出發開車）
        taiwan_mainland_cities = ['台北', '臺北', '新北', '桃園', '新竹', '苗栗', '台中', '臺中', 
                                '彰化', '南投', '雲林', '嘉義', '台南', '臺南', '高雄', '屏東',
                                '宜蘭', '花蓮', '台東', '臺東', '基隆']
        
        # 台灣離島（需要搭飛機）
        taiwan_islands = ['澎湖', '金門', '馬祖']
        
        # 台灣地區關鍵字（只包含台灣本地區域）
        taiwan_regions = list(TAIWAN_REGION_MAPPING.keys())
        
        # 檢查是否為台灣離島
        if any(island in destination for island in taiwan_islands):
            return False  # 離島視為非本島，使用飛機
            
        # 檢查是否為台灣本島地點
        is_taiwan_mainland = any(city in destination for city in taiwan_mainland_cities) or \
                           any(region in destination for region in taiwan_regions)
        
        # 檢查是否為中國城市
        is_china = any(city in destination for city in CHINA_KEYWORDS)
        
        # 如果是中國城市，視為國際
        if is_china:
            return False
            
        # 特殊地點判斷
        if destination in COORDINATES:
            coord_info = COORDINATES[destination]
            if coord_info['type'] == 'domestic':
                return True
            elif coord_info['type'] == 'island':
                return False  # 離島使用飛機
                
        return is_taiwan_mainland
        
    def _calculate_driving_distance(self, row, destination):
        """計算開車距離"""
        # 處理目的地名稱變化（例如：臺北/台北）
        destination_normalized = destination
        original_destination = destination
        
        # 檢查是否為台灣地區，需要對應到主要城市
        if destination in TAIWAN_REGION_MAPPING:
            destination_normalized = TAIWAN_REGION_MAPPING[destination]
            print(f"地區對應：{destination} -> {destination_normalized}")
        elif destination in INTERNATIONAL_CITY_MAPPING:
            # 國際城市不應該用開車，返回錯誤
            print(f"警告：國際城市{destination}被錯誤地傳入開車計算函數")
            return self._calculate_flight_distance(row, destination)
        
        if destination_normalized not in COORDINATES:
            # 嘗試簡體轉換
            if '臺' in destination_normalized:
                destination_normalized = destination_normalized.replace('臺', '台')
            elif '台' in destination_normalized:
                destination_normalized = destination_normalized.replace('台', '臺')
                
        if destination_normalized not in COORDINATES:
            # 如果還是找不到，記錄並跳過
            print(f"警告：找不到目的地座標 - {destination}")
            # 如果是複合地名，嘗試分解
            # 例如："台北市南港區" -> "南港"
            if original_destination != destination:
                # 已經做過處理
                pass
            else:
                # 嘗試分解地名
                for region in TAIWAN_REGION_MAPPING.keys():
                    if region in original_destination:
                        destination_normalized = TAIWAN_REGION_MAPPING[region]
                        print(f"地區分解對應：{original_destination} -> {region} -> {destination_normalized}")
                        destination = destination_normalized
                        break
                        
            if destination not in COORDINATES:
                print(f"警告：找不到目的地座標 - {original_destination}")
            
            return {
                '序號': row['序號'],
                '傳票編號': row['傳票編號'],
                '目的地': original_destination,
                '類型': '未知',
                '距離(km)': 0,
                '時間': 'N/A',
                '起點': {'name': '輔仁大學', 'lat': 25.0356, 'lon': 121.4320},
                '終點': {'name': original_destination, 'lat': 0, 'lon': 0},
                'route': None,
                'route_key': f"unknown-{original_destination}",
                'cached': False
            }
            
        destination = destination_normalized
        
        # 建立路線的唯一識別碼
        route_key = f"輔仁大學-{destination}-driving"
        
        # 檢查快取
        if route_key in self.route_cache:
            cached_data = self.route_cache[route_key]
            # 使用原始目的地名稱
            return {
                '序號': row['序號'],
                '傳票編號': row['傳票編號'],
                '目的地': f"{original_destination} ({cached_data['終點']['name']})",
                '類型': cached_data['類型'],
                '距離(km)': cached_data['距離(km)'],
                '時間': cached_data['時間'],
                '起點': cached_data['起點'],
                '終點': cached_data['終點'],
                'route': cached_data.get('route'),
                'route_key': route_key,
                'cached': True
            }
            
        try:
            # 使用OSRM API計算實際路線
            start = COORDINATES['輔仁大學']
            end = COORDINATES[destination]
            
            url = f"http://router.project-osrm.org/route/v1/driving/{start['lon']},{start['lat']};{end['lon']},{end['lat']}"
            params = {'overview': 'full', 'geometries': 'geojson'}
            
            response = requests.get(url, params=params, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                if data['code'] == 'Ok':
                    route = data['routes'][0]
                    distance_km = round(route['distance'] / 1000, 1)
                    duration_min = round(route['duration'] / 60)
                    
                    # 儲存到快取
                    self.route_cache[route_key] = {
                        '目的地': f"{original_destination} ({end['name']})",
                        '類型': '國內-開車',
                        '距離(km)': distance_km,
                        '時間': f"{duration_min}分鐘",
                        '起點': start,
                        '終點': end,
                        'route': route['geometry']
                    }
                    
                    return {
                        '序號': row['序號'],
                        '傳票編號': row['傳票編號'],
                        '目的地': f"{original_destination} ({end['name']})",
                        '類型': '國內-開車',
                        '距離(km)': distance_km,
                        '時間': f"{duration_min}分鐘",
                        '起點': start,
                        '終點': end,
                        'route': route['geometry'],
                        'route_key': route_key,
                        'cached': False
                    }
        except:
            pass
            
        # 如果API失敗，使用直線距離估算
        start = COORDINATES['輔仁大學']
        end = COORDINATES[destination]
        start_point = (start['lat'], start['lon'])
        end_point = (end['lat'], end['lon'])
        distance_km = round(geodesic(start_point, end_point).kilometers * 1.4, 1)
        duration_min = round(distance_km * 1.5)
        
        # 儲存到快取
        self.route_cache[route_key] = {
            '目的地': f"{original_destination} ({end['name']})",
            '類型': '國內-開車',
            '距離(km)': distance_km,
            '時間': f"{duration_min}分鐘",
            '起點': start,
            '終點': end,
            'route': None
        }
        
        return {
            '序號': row['序號'],
            '傳票編號': row['傳票編號'],
            '目的地': f"{destination} ({end['name']})",
            '類型': '國內-開車',
            '距離(km)': distance_km,
            '時間': f"{duration_min}分鐘",
            '起點': start,
            '終點': end,
            'route': None,
            'route_key': route_key,
            'cached': False
        }
        
    def _calculate_island_flight_distance(self, row, destination):
        """計算台灣離島航班距離（從松山機場或桃園機場）"""
        original_destination = destination
        
        if destination not in COORDINATES:
            return {
                '序號': row['序號'],
                '傳票編號': row['傳票編號'],
                '目的地': original_destination,
                '類型': '離島-未知',
                '距離(km)': 0,
                '時間': 'N/A',
                '起點': {'name': '松山機場', 'lat': 25.0694, 'lon': 121.5526},
                '終點': {'name': original_destination, 'lat': 0, 'lon': 0},
                'route': None,
                'route_key': f"unknown-{original_destination}",
                'cached': False
            }
        
        # 根據離島選擇合適的起飛機場
        if destination in ['澎湖', '金門', '馬祖'] or (destination in COORDINATES and COORDINATES[destination].get('type') == 'island'):
            # 離島通常從松山機場起飛
            start = COORDINATES['松山機場']
            flight_type = '國內-離島航班'
        else:
            start = COORDINATES['桃園機場']
            flight_type = '國內-航班'
        
        # 建立路線的唯一識別碼
        route_key = f"{start['name']}-{destination}-island-flight"
        
        # 檢查快取
        if route_key in self.route_cache:
            cached_data = self.route_cache[route_key]
            return {
                '序號': row['序號'],
                '傳票編號': row['傳票編號'],
                '目的地': f"{original_destination} ({cached_data['終點']['name']})",
                '類型': cached_data['類型'],
                '距離(km)': cached_data['距離(km)'],
                '時間': cached_data['時間'],
                '起點': cached_data['起點'],
                '終點': cached_data['終點'],
                'route': None,
                'route_key': route_key,
                'cached': True
            }
            
        # 計算離島航班距離
        end = COORDINATES[destination]
        
        # 計算大圓距離
        start_point = (start['lat'], start['lon'])
        end_point = (end['lat'], end['lon'])
        distance_km = round(geodesic(start_point, end_point).kilometers, 1)
        
        # 估算飛行時間（離島航班通常較短）
        flight_hours = distance_km / 500 + 0.5  # 離島航班速度較慢
        
        # 儲存到快取
        self.route_cache[route_key] = {
            '目的地': f"{destination} ({end['name']})",
            '類型': flight_type,
            '距離(km)': distance_km,
            '時間': f"{flight_hours:.1f}小時",
            '起點': start,
            '終點': end
        }
        
        return {
            '序號': row['序號'],
            '傳票編號': row['傳票編號'],
            '目的地': f"{original_destination} ({end['name']})",
            '類型': flight_type,
            '距離(km)': distance_km,
            '時間': f"{flight_hours:.1f}小時",
            '起點': start,
            '終點': end,
            'route': None,
            'route_key': route_key,
            'cached': False
        }
        
    def _calculate_flight_distance(self, row, destination):
        """計算國際飛行距離（統一從桃園機場出發）"""
        original_destination = destination
        
        # 檢查是否為需要對應的國際城市
        if destination in INTERNATIONAL_CITY_MAPPING:
            destination = INTERNATIONAL_CITY_MAPPING[destination]
            print(f"國際城市對應：{original_destination} -> {destination}")
        elif destination in TAIWAN_REGION_MAPPING:
            # 台灣地區不應該用飛機，返回開車計算
            print(f"警告：台灣地區{destination}被錯誤地傳入飛機計算函數")
            return self._calculate_driving_distance(row, destination)
        
        # 特別處理中國城市
        if any(city in original_destination for city in CHINA_KEYWORDS):
            # 如果是中國城市但不在座標中，嘗試找最近的機場
            if destination not in COORDINATES:
                # 檢查是否有對應的機場
                for china_city, airport in INTERNATIONAL_CITY_MAPPING.items():
                    if china_city in original_destination:
                        destination = airport
                        print(f"中國城市對應機場：{original_destination} -> {destination}")
                        break
        
        if destination not in COORDINATES:
            # 如果找不到座標，記錄並跳過
            print(f"警告：找不到國際目的地座標 - {original_destination}")
            return {
                '序號': row['序號'],
                '傳票編號': row['傳票編號'],
                '目的地': original_destination,
                '類型': '國際-未知',
                '距離(km)': 0,
                '時間': 'N/A',
                '起點': {'name': '桃園機場', 'lat': 25.0777, 'lon': 121.2325},
                '終點': {'name': original_destination, 'lat': 0, 'lon': 0},
                'route': None,
                'route_key': f"unknown-{original_destination}",
                'cached': False
            }
        
        # 建立路線的唯一識別碼
        route_key = f"桃園機場-{destination}-flying"
        
        # 檢查快取
        if route_key in self.route_cache:
            cached_data = self.route_cache[route_key]
            # 使用原始目的地名稱
            return {
                '序號': row['序號'],
                '傳票編號': row['傳票編號'],
                '目的地': f"{original_destination} ({cached_data['終點']['name']})",
                '類型': cached_data['類型'],
                '距離(km)': cached_data['距離(km)'],
                '時間': cached_data['時間'],
                '起點': cached_data['起點'],
                '終點': cached_data['終點'],
                'route': None,
                'route_key': route_key,
                'cached': True
            }
            
        # 從桃園機場飛往目的地
        start = COORDINATES['桃園機場']
        end = COORDINATES[destination]
        
        # 計算大圓距離
        start_point = (start['lat'], start['lon'])
        end_point = (end['lat'], end['lon'])
        distance_km = round(geodesic(start_point, end_point).kilometers, 1)
        
        # 估算飛行時間（包含起降）
        flight_hours = distance_km / 800 + 0.5
        
        # 儲存到快取
        self.route_cache[route_key] = {
            '目的地': f"{destination} ({end['name']})",
            '類型': '國際-飛行',
            '距離(km)': distance_km,
            '時間': f"{flight_hours:.1f}小時",
            '起點': start,
            '終點': end
        }
        
        return {
            '序號': row['序號'],
            '傳票編號': row['傳票編號'],
            '目的地': f"{original_destination} ({end['name']})",
            '類型': '國際-飛行',
            '距離(km)': distance_km,
            '時間': f"{flight_hours:.1f}小時",
            '起點': start,
            '終點': end,
            'route': None,
            'route_key': route_key,
            'cached': False
        }
        
    def _add_result_to_tree(self, result):
        """添加結果到Treeview"""
        values = (
            result['序號'],
            result['傳票編號'],
            result['目的地'],
            result['類型'],
            result['距離(km)'],
            result['時間'],
            '待產生',
            '✓' if result.get('cached') else ''
        )
        self.tree.insert('', 'end', values=values)
        
    def generate_maps(self):
        """產生地圖"""
        if not self.results:
            messagebox.showerror("錯誤", "請先計算距離")
            return
            
        # 在新執行緒中執行
        thread = threading.Thread(target=self._generate_maps_thread)
        thread.start()
        
    def _generate_maps_thread(self):
        """產生地圖的執行緒函數 - 優化版本"""
        try:
            # 確保輸出目錄存在
            map_dir = os.path.join(self.output_dir.get(), 'maps')
            os.makedirs(map_dir, exist_ok=True)
            
            # 只處理唯一路線，不是每個結果
            total_unique_routes = len(self.unique_routes)
            cache_hit = 0
            new_maps = 0
            
            self.status_text.set(f"準備產生 {total_unique_routes} 個唯一路線的地圖...")
            
            # 只為每個唯一路線產生一次地圖
            for idx, (route_key, route_info) in enumerate(self.unique_routes.items()):
                self.status_text.set(f"正在處理唯一路線 {idx+1}/{total_unique_routes}...")
                self.progress.set((idx + 1) / total_unique_routes * 100)
                
                # 檢查是否已有快取的地圖
                if route_key in self.map_cache and os.path.exists(self.map_cache[route_key]):
                    # 使用已存在的地圖檔案
                    map_file = self.map_cache[route_key]
                    cache_hit += 1
                    self.status_text.set(f"使用快取地圖: {route_key}")
                else:
                    # 產生新地圖（使用第一個結果作為範本）
                    sample_result = route_info['results'][0]
                    map_file = self._create_map(sample_result, map_dir, route_key)
                    
                    if map_file:
                        # 截圖
                        screenshot = self._capture_map_screenshot(map_file)
                        if screenshot:
                            map_file = screenshot
                            
                        # 儲存到快取
                        self.map_cache[route_key] = map_file
                        new_maps += 1
                        self.status_text.set(f"已產生新地圖: {route_key}")
                
                # 將地圖檔案分配給所有使用此路線的結果
                if map_file:
                    for result in route_info['results']:
                        result['地圖檔案'] = map_file
                        # 更新UI顯示
                        self._update_tree_for_result(result)
                    
            # 儲存快取
            self.save_cache()
            self.update_cache_info()
            
            # 計算效率提升
            total_results = len(self.results)
            efficiency_ratio = round((1 - total_unique_routes / total_results) * 100, 1) if total_results > 0 else 0
            
            self.status_text.set(
                f"地圖產生完成！共 {total_results} 筆資料，"
                f"僅需產生 {total_unique_routes} 個地圖，"
                f"效率提升 {efficiency_ratio}%，"
                f"快取 {cache_hit} 個，新建 {new_maps} 個"
            )
            self.progress.set(100)
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("錯誤", f"產生地圖失敗: {str(e)}"))
            self.status_text.set("產生地圖失敗")
            
    def _update_tree_for_result(self, result):
        """更新特定結果在 TreeView 中的顯示"""
        items = self.tree.get_children()
        for item in items:
            values = list(self.tree.item(item)['values'])
            if values[0] == result['序號'] and values[1] == result['傳票編號']:
                values[-2] = '已產生'  # 更新地圖狀態
                self.tree.item(item, values=values)
                break
                
    def _create_map(self, result, output_dir, name):
        """創建地圖（包含智能縮放）"""
        try:
            # 取得起終點座標
            start_lat = result['起點']['lat']
            start_lon = result['起點']['lon']
            end_lat = result['終點']['lat']
            end_lon = result['終點']['lon']
            
            # 計算中心點
            center_lat = (start_lat + end_lat) / 2
            center_lon = (start_lon + end_lon) / 2
            
            # 根據距離決定縮放級別
            distance = result.get('距離(km)', 0)
            zoom_level = calculate_zoom_level(distance)
            
            # 創建地圖
            m = folium.Map(
                location=[center_lat, center_lon], 
                zoom_start=zoom_level,
                tiles='OpenStreetMap',
                prefer_canvas=True
            )
            
            # 添加起點標記（綠色）
            folium.Marker(
                [start_lat, start_lon],
                popup=f"<b>起點</b><br>{result['起點']['name']}",
                icon=folium.Icon(color='green', icon='play', prefix='fa'),
                tooltip="起點"
            ).add_to(m)
            
            # 添加終點標記（紅色）
            folium.Marker(
                [end_lat, end_lon],
                popup=f"<b>終點</b><br>{result['終點']['name']}",
                icon=folium.Icon(color='red', icon='stop', prefix='fa'),
                tooltip="終點"
            ).add_to(m)
            
            # 添加路線
            if result['類型'] == '國內-開車' and result.get('route'):
                # 開車路線（詳細路徑）
                route_coords = [[coord[1], coord[0]] for coord in result['route']['coordinates']]
                
                # 主路線
                folium.PolyLine(
                    route_coords,
                    color='#0066CC',
                    weight=6,
                    opacity=0.8,
                    smooth_factor=1
                ).add_to(m)
                
                # 路線邊框（讓路線更明顯）
                folium.PolyLine(
                    route_coords,
                    color='#000033',
                    weight=8,
                    opacity=0.4,
                    smooth_factor=1
                ).add_to(m)
                
                # 使用路線邊界來調整地圖視野
                if len(route_coords) > 0:
                    lats = [coord[0] for coord in route_coords]
                    lons = [coord[1] for coord in route_coords]
                    sw = [min(lats), min(lons)]
                    ne = [max(lats), max(lons)]
                    
                    # 根據距離調整padding
                    if distance < 20:
                        padding = 0.02  # 短距離用較小的padding
                    elif distance < 50:
                        padding = 0.05
                    else:
                        padding = 0.1
                        
                    m.fit_bounds([sw, ne], padding=(padding, padding))
                    
            else:
                # 飛行路線（曲線）
                # 添加飛行路線的曲線效果（貝茲曲線）
                import numpy as np
                t = np.linspace(0, 1, 50)
                
                # 計算控制點（讓路線呈現弧形）
                control_lat = center_lat + (distance / 10000) * 5  # 根據距離調整弧度
                control_lon = center_lon
                
                # 二次貝茲曲線
                curve_lats = (1-t)**2 * start_lat + 2*(1-t)*t * control_lat + t**2 * end_lat
                curve_lons = (1-t)**2 * start_lon + 2*(1-t)*t * control_lon + t**2 * end_lon
                
                curve_coords = [[lat, lon] for lat, lon in zip(curve_lats, curve_lons)]
                
                # 飛行路線（橘色曲線）
                folium.PolyLine(
                    curve_coords,
                    color='#FF6600',
                    weight=4,
                    opacity=0.8
                ).add_to(m)
                
                # 虛線直線（參考用）
                folium.PolyLine(
                    [[start_lat, start_lon], [end_lat, end_lon]],
                    color='#CC0000',
                    weight=2,
                    opacity=0.4,
                    dash_array='10, 5'
                ).add_to(m)
                
                # 確保兩個機場都在視野內
                sw = [min(start_lat, end_lat), min(start_lon, end_lon)]
                ne = [max(start_lat, end_lat), max(start_lon, end_lon)]
                
                # 國際航線需要更大的padding
                padding = 0.2 if distance > 1000 else 0.1
                m.fit_bounds([sw, ne], padding=(padding, padding))
            
            # 添加距離標記（在路線中點）
            folium.Marker(
                [center_lat, center_lon],
                icon=folium.DivIcon(
                    html=f"""
                    <div style="background-color: white; 
                                border: 2px solid black; 
                                border-radius: 5px; 
                                padding: 5px;
                                font-weight: bold;
                                font-size: 12px;">
                        {distance} km
                    </div>
                    """
                )
            ).add_to(m)
            
            # 添加資訊框（固定在右上角）- 不顯示特定傳票編號
            info_html = f"""
            <div style='position: fixed; 
                        top: 10px; right: 10px; 
                        background: rgba(255,255,255,0.95); 
                        padding: 15px;
                        border: 2px solid #333;
                        border-radius: 8px;
                        font-family: "Microsoft JhengHei", Arial;
                        box-shadow: 0 2px 6px rgba(0,0,0,0.3);
                        z-index: 9999;
                        max-width: 300px;'>
                <h3 style='margin: 0 0 10px 0; color: #333;'>路線資訊</h3>
                <table style='font-size: 14px;'>
                    <tr><td style='padding: 3px;'><b>起點:</b></td><td>{result['起點']['name']}</td></tr>
                    <tr><td style='padding: 3px;'><b>終點:</b></td><td>{result['終點']['name']}</td></tr>
                    <tr><td style='padding: 3px;'><b>距離:</b></td><td style='color: #0066CC; font-weight: bold;'>{distance} km</td></tr>
                    <tr><td style='padding: 3px;'><b>時間:</b></td><td>{result.get('時間', 'N/A')}</td></tr>
                    <tr><td style='padding: 3px;'><b>類型:</b></td><td>{result.get('類型', 'N/A')}</td></tr>
                </table>
            </div>
            """
            m.get_root().html.add_child(folium.Element(info_html))
            
            # 儲存地圖
            filename = f"map_{name}.html"
            filepath = os.path.join(output_dir, filename)
            m.save(filepath)
            
            return filepath
            
        except Exception as e:
            print(f"創建地圖失敗: {e}")
            return None
            
    def _capture_map_screenshot(self, html_file):
        """截取地圖截圖"""
        try:
            # 設置Chrome選項
            options = Options()
            options.add_argument('--headless')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--window-size=1200,800')
            
            # 創建瀏覽器實例
            driver = webdriver.Chrome(options=options)
            
            # 載入地圖
            driver.get(f"file://{html_file}")
            
            # 等待地圖載入
            time.sleep(3)
            
            # 截圖
            screenshot_path = html_file.replace('.html', '.png')
            driver.save_screenshot(screenshot_path)
            
            driver.quit()
            
            return screenshot_path
            
        except Exception as e:
            print(f"截圖失敗: {e}")
            # 如果Chrome截圖失敗，返回None
            return None
            
    def _update_tree_item(self, index, status):
        """更新Treeview項目"""
        items = self.tree.get_children()
        if index < len(items):
            item = items[index]
            values = list(self.tree.item(item)['values'])
            values[-2] = status  # 地圖狀態
            self.tree.item(item, values=values)
            
    def export_excel(self):
        """匯出Excel"""
        if not self.results:
            messagebox.showerror("錯誤", "沒有可匯出的資料")
            return
            
        try:
            self.status_text.set("正在匯出Excel...")
            
            # 建立DataFrame
            export_data = []
            for result in self.results:
                carbon = calculate_carbon_emission(result['類型'], result['距離(km)'])
                export_data.append({
                    '序號': result['序號'],
                    '傳票編號': result['傳票編號'],
                    '目的地': result['目的地'],
                    '類型': result['類型'],
                    '距離(km)': result['距離(km)'],
                    '時間': result['時間'],
                    '起點': result['起點']['name'],
                    '終點': result['終點']['name'],
                    '交通方式': carbon['交通方式'],
                    '碳排係數': carbon['碳排係數'],
                    '碳排放量(kg CO2e)': carbon['碳排放量(kg CO2e)'],
                })
                
            df = pd.DataFrame(export_data)
            
            # 儲存Excel
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(self.output_dir.get(), f'出差距離計算結果_{timestamp}.xlsx')
            
            # 使用ExcelWriter來插入圖片
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='距離計算結果', index=False)
                
                # 取得工作表
                workbook = writer.book
                worksheet = writer.sheets['距離計算結果']
                
                # 調整欄寬
                column_widths = {'A': 10, 'B': 15, 'C': 30, 'D': 15, 'E': 12, 'F': 12, 'G': 20, 'H': 20,
                                 'I': 18, 'J': 12, 'K': 20}
                for column, width in column_widths.items():
                    worksheet.column_dimensions[column].width = width
                    
                # 插入地圖截圖（如果有的話）
                if any('地圖檔案' in r for r in self.results):
                    # 新增一個工作表放地圖
                    ws_maps = workbook.create_sheet('地圖')
                    
                    # 使用已經整理好的唯一路線資訊
                    row = 1
                    for route_key, route_info in self.unique_routes.items():
                        results = route_info['results']
                        result = results[0]  # 取第一個結果作為範本
                        
                        if '地圖檔案' in result and result['地圖檔案'] and os.path.exists(result['地圖檔案']):
                            # 添加標題
                            ws_maps.cell(row=row, column=1, value=f"路線: {result['起點']['name']} → {result['終點']['name']}")
                            ws_maps.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True, size=14)
                            row += 1
                            
                            # 列出使用此路線的所有傳票
                            vouchers = [str(r['傳票編號']) for r in results]  # 確保轉換為字串
                            ws_maps.cell(row=row, column=1, value=f"傳票編號: {', '.join(vouchers)}")
                            ws_maps.cell(row=row, column=2, value=f"(共 {len(vouchers)} 筆)")
                            row += 1
                            
                            # 列出距離和時間
                            ws_maps.cell(row=row, column=1, value=f"距離: {result['距離(km)']} km | 時間: {result['時間']}")
                            row += 1
                            
                            # 插入圖片
                            try:
                                if result['地圖檔案'].endswith('.png'):
                                    img = XLImage(result['地圖檔案'])
                                    img.width = 600
                                    img.height = 400
                                    ws_maps.add_image(img, f'A{row}')
                                    row += 25  # 預留圖片空間
                                else:
                                    ws_maps.cell(row=row, column=1, value=f"地圖檔案: {result['地圖檔案']}")
                                    row += 2
                            except Exception as e:
                                print(f"插入圖片失敗: {e}")
                                ws_maps.cell(row=row, column=1, value=f"地圖檔案: {result['地圖檔案']}")
                                row += 2
                            
                            # 添加分隔線
                            row += 2

                # ========== 碳排放總覽工作表 ==========
                ws_carbon = workbook.create_sheet('碳排放總覽')

                # 標題樣式
                title_font = openpyxl.styles.Font(bold=True, size=14)
                header_font = openpyxl.styles.Font(bold=True, size=11)
                header_fill = openpyxl.styles.PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

                # 計算碳排放資料
                carbon_by_mode = defaultdict(lambda: {'trips': 0, 'distance': 0, 'emission': 0})
                total_emission = 0
                total_trips = 0
                for r in self.results:
                    dist = r.get('距離(km)', 0)
                    if dist <= 0:
                        continue
                    c = calculate_carbon_emission(r['類型'], dist)
                    mode = c['交通方式']
                    if mode == 'N/A':
                        continue
                    carbon_by_mode[mode]['trips'] += 1
                    carbon_by_mode[mode]['distance'] += dist
                    carbon_by_mode[mode]['emission'] += c['碳排放量(kg CO2e)']
                    total_emission += c['碳排放量(kg CO2e)']
                    total_trips += 1

                # 總覽標題
                ws_carbon.cell(row=1, column=1, value='碳排放總覽').font = title_font
                ws_carbon.cell(row=2, column=1, value=f'產生日期：{datetime.now().strftime("%Y-%m-%d %H:%M")}')

                # 總計區塊
                ws_carbon.cell(row=4, column=1, value='總碳排放量(kg CO2e)').font = header_font
                ws_carbon.cell(row=4, column=2, value=round(total_emission, 3))
                ws_carbon.cell(row=5, column=1, value='有效出差筆數').font = header_font
                ws_carbon.cell(row=5, column=2, value=total_trips)
                ws_carbon.cell(row=6, column=1, value='平均每趟碳排(kg CO2e)').font = header_font
                ws_carbon.cell(row=6, column=2, value=round(total_emission / total_trips, 3) if total_trips > 0 else 0)

                # 各交通方式明細表
                ws_carbon.cell(row=8, column=1, value='各交通方式碳排明細').font = title_font
                detail_headers = ['交通方式', '出差筆數', '總距離(km)', '碳排係數(kg/km)', '碳排放量(kg CO2e)', '佔比(%)']
                for col_idx, h in enumerate(detail_headers, 1):
                    cell = ws_carbon.cell(row=9, column=col_idx, value=h)
                    cell.font = header_font
                    cell.fill = header_fill

                detail_row = 10
                for mode, data in sorted(carbon_by_mode.items(), key=lambda x: x[1]['emission'], reverse=True):
                    factor = CARBON_EMISSION_FACTORS.get(mode, 0)
                    pct = round(data['emission'] / total_emission * 100, 1) if total_emission > 0 else 0
                    ws_carbon.cell(row=detail_row, column=1, value=mode)
                    ws_carbon.cell(row=detail_row, column=2, value=data['trips'])
                    ws_carbon.cell(row=detail_row, column=3, value=round(data['distance'], 1))
                    ws_carbon.cell(row=detail_row, column=4, value=factor)
                    ws_carbon.cell(row=detail_row, column=5, value=round(data['emission'], 3))
                    ws_carbon.cell(row=detail_row, column=6, value=pct)
                    detail_row += 1

                # 合計列
                ws_carbon.cell(row=detail_row, column=1, value='合計').font = header_font
                ws_carbon.cell(row=detail_row, column=2, value=total_trips).font = header_font
                total_dist = sum(d['distance'] for d in carbon_by_mode.values())
                ws_carbon.cell(row=detail_row, column=3, value=round(total_dist, 1)).font = header_font
                ws_carbon.cell(row=detail_row, column=5, value=round(total_emission, 3)).font = header_font
                ws_carbon.cell(row=detail_row, column=6, value=100.0).font = header_font

                # 調整碳排放工作表欄寬
                carbon_col_widths = {'A': 20, 'B': 15, 'C': 15, 'D': 18, 'E': 22, 'F': 12}
                for col, w in carbon_col_widths.items():
                    ws_carbon.column_dimensions[col].width = w
                # ========== 碳排放總覽結束 ==========

            messagebox.showinfo("成功", f"Excel已匯出至:\n{output_file}")
            self.status_text.set("匯出完成")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"匯出失敗: {str(e)}")
            self.status_text.set("匯出失敗")
            
    def run(self):
        """執行GUI"""
        # 確保視窗顯示
        self.root.update()
        self.root.deiconify()
        print("GUI視窗已啟動...")
        self.root.mainloop()


# 主程式
if __name__ == "__main__":
    # 確保必要的套件已安裝
    required_packages = {
        'pandas': 'pandas',
        'folium': 'folium',
        'selenium': 'selenium',
        'PIL': 'pillow',
        'geopy': 'geopy',
        'openpyxl': 'openpyxl',
        'requests': 'requests'
    }
    
    missing_packages = []
    for import_name, install_name in required_packages.items():
        try:
            __import__(import_name)
        except ImportError:
            missing_packages.append(install_name)
            
    if missing_packages:
        print("請先安裝以下套件:")
        print(f"pip install {' '.join(missing_packages)}")
        sys.exit(1)
    
    print("正在啟動出差距離計算與地圖產生工具...")
    
    try:
        # 啟動應用程式
        app = TravelDistanceCalculator()
        app.run()
    except Exception as e:
        print(f"啟動失敗: {e}")
        import traceback
        traceback.print_exc()
